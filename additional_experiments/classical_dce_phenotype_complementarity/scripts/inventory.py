#!/usr/bin/env python3
"""Fail-closed, aggregate-only inventory of the classical DCE workbook.

The command validates the locked radiomics workbook, clinical workbook, and MRI
fold manifest before writing any output.  Patient identifiers are used in memory
only for exact set intersection; no output contains identifier values.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import numpy as np
import pandas as pd
from openpyxl import load_workbook


EXPERIMENT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG = EXPERIMENT_ROOT / "configs" / "experiment.json"

ID_COLUMN = "CLINICAL-TRIAL-SUBJECT-ID"
EXPECTED_RAD_ROWS = 384
EXPECTED_CLINICAL_ROWS = 985
EXPECTED_MRI_PATIENTS = 808
EXPECTED_MRI_OVERLAP = 375

EXPECTED_ABSOLUTE: dict[str, tuple[str, ...]] = {
    "FTV": (
        "VOLUME_TUM_BLU_V10",
        "VOLUME_TUM_BLU_V20",
        "VOLUME_TUM_BLU_V30",
        "VOLUME_TUM_BLU_V40",
    ),
    "SPH": (
        "SPHERICITY_T0",
        "SPHERICITY_T1",
        "SPHERICITY_T2",
        "SPHERICITY_T3",
    ),
    "LD": ("LD_T0", "LD_T1", "LD_T2", "LD_T3"),
    "BPE": (
        "BPE_5slice_mean_T0",
        "BPE_5slice_mean_T1",
        "BPE_5slice_mean_T2",
        "BPE_5slice_mean_T3",
    ),
}

EXPECTED_PERCENT_CHANGE: dict[str, tuple[str, ...]] = {
    "FTV": ("FTV_pch_T0_T1", "FTV_pch_T0_T2", "FTV_pch_T0_T3"),
    "SPH": (
        "Sphericity_pch_T0_T1",
        "Sphericity_pch_T0_T2",
        "Sphericity_pch_T0_T3",
    ),
    "LD": ("LD_pch_T0_T1", "LD_pch_T0_T2", "LD_pch_T0_T3"),
    "BPE": ("BPE_pch_T0_T1", "BPE_pch_T0_T2", "BPE_pch_T0_T3"),
}

EXPECTED_CLINICAL_COLUMNS = (
    "Patient_ID",
    "Arm",
    "HR",
    "HER2",
    "MP",
    "pCR",
    "Age_at_Screening",
    "Race",
    "menopausal_status",
    "ethnicity",
)

FAMILY_UNIT = {
    "FTV": "cc",
    "SPH": "dimensionless",
    "LD": "not declared in workbook",
    "BPE": "native mean-PE scale; absolute unit not declared in workbook",
}

FAMILY_DEFINITION = {
    "FTV": "Functional tumor volume satisfying the I-SPY2 PE/SER analysis criteria.",
    "SPH": "Sphericity derived from the three-dimensional FTV tumor mask.",
    "LD": "Imaging-reported longest tumor diameter.",
    "BPE": "Contralateral central-five-slice fibroglandular mean early enhancement.",
}


class InventoryError(RuntimeError):
    """Raised when a locked input or privacy/schema contract is violated."""


@dataclass(frozen=True)
class ColumnSpec:
    column: str
    family: str
    visit: str
    unit: str
    role: str
    definition: str
    first_usable_timing: str
    independent_measurement: bool
    baseline_column: str | None = None
    endpoint_column: str | None = None


@dataclass(frozen=True)
class WorkbookAudit:
    path: Path
    sheet: str
    sha256: str
    rows: int
    columns: int
    used_range: str
    formula_cells: int
    duplicate_patient_rows: int
    duplicate_full_rows: int


@dataclass(frozen=True)
class ClinicalAudit:
    path: Path
    sheet: str
    sha256: str
    rows: int
    columns: int
    radiomics_overlap: int
    target_summary: tuple[dict[str, Any], ...]
    subtype_summary: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class MriCoverage:
    path: Path
    sha256: str
    cohort_patients: int
    radiomics_overlap: int
    workbook_only: int
    mri_only: int
    matched_workbook_mask: np.ndarray


@dataclass(frozen=True)
class InventoryResult:
    inventory_path: Path
    missingness_path: Path
    report_path: Path
    inventory_rows: int
    missingness_rows: int
    workbook_patients: int
    mri_overlap: int


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _load_config(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise InventoryError(f"config is not a file: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise InventoryError(f"cannot read config: {path}") from exc
    if not isinstance(payload, dict) or not isinstance(payload.get("source"), dict):
        raise InventoryError("config must contain a source object")
    return payload


def _required_path(source: Mapping[str, Any], key: str) -> Path:
    value = source.get(key)
    if not isinstance(value, str) or not value.strip():
        raise InventoryError(f"config source.{key} must be a non-empty path")
    path = Path(value).expanduser()
    if not path.is_file():
        raise InventoryError(f"configured source is not a file: {path}")
    return path


def _required_text(source: Mapping[str, Any], key: str) -> str:
    value = source.get(key)
    if not isinstance(value, str) or not value.strip():
        raise InventoryError(f"config source.{key} must be non-empty text")
    return value.strip()


def _validate_sha(path: Path, expected: str, label: str) -> str:
    if not re.fullmatch(r"[0-9a-f]{64}", expected):
        raise InventoryError(f"configured {label} SHA-256 is malformed")
    observed = _sha256(path)
    if observed != expected:
        raise InventoryError(
            f"{label} SHA-256 mismatch (expected {expected}, observed {observed})"
        )
    return observed


def _normalise_six_digit_ids(
    values: pd.Series,
    *,
    label: str,
    expected_rows: int | None = None,
) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce")
    invalid_numeric = int(numeric.isna().sum())
    finite = np.isfinite(numeric.to_numpy(dtype=float, na_value=np.nan))
    non_integer = int(
        np.sum(finite & (numeric.to_numpy(dtype=float, na_value=np.nan) % 1 != 0))
    )
    if invalid_numeric or non_integer:
        raise InventoryError(
            f"{label} has invalid ID cells: nonnumeric={invalid_numeric}, "
            f"noninteger={non_integer}"
        )
    canonical = numeric.astype("int64").astype(str)
    bad_format = int((~canonical.str.fullmatch(r"\d{6}")).sum())
    duplicates = int(canonical.duplicated().sum())
    if bad_format or duplicates:
        raise InventoryError(
            f"{label} ID contract failed: bad_six_digit_format={bad_format}, "
            f"duplicates={duplicates}"
        )
    if expected_rows is not None and len(canonical) != expected_rows:
        raise InventoryError(
            f"{label} row count mismatch: expected {expected_rows}, observed {len(canonical)}"
        )
    return canonical


def _build_column_specs(config: Mapping[str, Any]) -> tuple[ColumnSpec, ...]:
    configured = config.get("feature_families")
    if not isinstance(configured, dict):
        raise InventoryError("config.feature_families must be an object")
    normalised = {
        str(family): tuple(str(column) for column in columns)
        for family, columns in configured.items()
        if isinstance(columns, list)
    }
    if normalised != EXPECTED_ABSOLUTE:
        raise InventoryError(
            "config.feature_families does not exactly match the locked four-family schema"
        )

    specs: list[ColumnSpec] = [
        ColumnSpec(
            column=ID_COLUMN,
            family="ID",
            visit="not_applicable",
            unit="not_applicable",
            role="patient_identifier",
            definition="Unique six-digit Clinical Trial Subject identifier.",
            first_usable_timing="join_only",
            independent_measurement=False,
        )
    ]
    visits = ("T0", "T1", "T2", "T3")
    for family, columns in EXPECTED_ABSOLUTE.items():
        for visit, column in zip(visits, columns, strict=True):
            specs.append(
                ColumnSpec(
                    column=column,
                    family=family,
                    visit=visit,
                    unit=FAMILY_UNIT[family],
                    role="absolute_measurement",
                    definition=FAMILY_DEFINITION[family],
                    first_usable_timing=visit,
                    independent_measurement=True,
                )
            )
    for family, columns in EXPECTED_PERCENT_CHANGE.items():
        baseline = EXPECTED_ABSOLUTE[family][0]
        for endpoint_index, column in enumerate(columns, start=1):
            endpoint = EXPECTED_ABSOLUTE[family][endpoint_index]
            end_visit = visits[endpoint_index]
            specs.append(
                ColumnSpec(
                    column=column,
                    family=family,
                    visit=f"T0→{end_visit}",
                    unit="%",
                    role="derived_baseline_percent_change",
                    definition="100 × (X_Tk − X_T0) / X_T0; materialized from absolute values.",
                    first_usable_timing=end_visit,
                    independent_measurement=False,
                    baseline_column=baseline,
                    endpoint_column=endpoint,
                )
            )
    return tuple(specs)


def _read_excel_metadata(path: Path, expected_sheet: str) -> tuple[str, int]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != [expected_sheet]:
            raise InventoryError(
                "workbook sheet contract failed: expected exactly one configured sheet"
            )
        worksheet = workbook[expected_sheet]
        if worksheet.sheet_state != "visible":
            raise InventoryError("configured worksheet must be visible")
        formula_cells = sum(
            1
            for row in worksheet.iter_rows()
            for cell in row
            if getattr(cell, "data_type", None) == "f"
        )
        return worksheet.calculate_dimension(), formula_cells
    finally:
        workbook.close()


def _read_radiomics_workbook(
    config: Mapping[str, Any], specs: Sequence[ColumnSpec]
) -> tuple[pd.DataFrame, pd.Series, WorkbookAudit, float]:
    source = config["source"]
    path = _required_path(source, "radiomics_workbook")
    sheet = _required_text(source, "radiomics_sheet")
    digest = _validate_sha(
        path, _required_text(source, "radiomics_sha256"), "radiomics workbook"
    )

    excel = pd.ExcelFile(path, engine="openpyxl")
    if excel.sheet_names != [sheet]:
        raise InventoryError(
            "radiomics workbook sheet contract failed: expected exactly one configured sheet"
        )
    frame = pd.read_excel(excel, sheet_name=sheet)
    expected_columns = [spec.column for spec in specs]
    observed_columns = [str(column) for column in frame.columns]
    if observed_columns != expected_columns:
        missing = sorted(set(expected_columns) - set(observed_columns))
        unexpected = sorted(set(observed_columns) - set(expected_columns))
        order_only = not missing and not unexpected
        raise InventoryError(
            "radiomics schema mismatch: "
            f"missing_columns={missing}, unexpected_columns={unexpected}, "
            f"order_only_mismatch={order_only}"
        )
    if frame.shape != (EXPECTED_RAD_ROWS, len(expected_columns)):
        raise InventoryError(
            "radiomics shape mismatch: "
            f"expected {(EXPECTED_RAD_ROWS, len(expected_columns))}, observed {frame.shape}"
        )

    ids = _normalise_six_digit_ids(
        frame[ID_COLUMN], label="radiomics workbook", expected_rows=EXPECTED_RAD_ROWS
    )
    feature_columns = expected_columns[1:]
    numeric = frame[feature_columns].apply(pd.to_numeric, errors="coerce")
    newly_invalid = int(
        ((frame[feature_columns].notna()) & numeric.isna()).to_numpy().sum()
    )
    nonfinite = int(np.isinf(numeric.to_numpy(dtype=float)).sum())
    if newly_invalid or nonfinite:
        raise InventoryError(
            "radiomics numeric contract failed: "
            f"nonnumeric_cells={newly_invalid}, infinite_cells={nonfinite}"
        )
    frame = frame.copy()
    frame[feature_columns] = numeric
    missing_cells = int(frame[expected_columns].isna().to_numpy().sum())
    if missing_cells:
        raise InventoryError(
            f"locked complete-case workbook unexpectedly has {missing_cells} missing cells"
        )
    duplicate_rows = int(ids.duplicated().sum())
    duplicate_full_rows = int(frame.duplicated().sum())
    if duplicate_rows or duplicate_full_rows:
        raise InventoryError(
            "radiomics duplicate contract failed: "
            f"duplicate_patient_rows={duplicate_rows}, duplicate_full_rows={duplicate_full_rows}"
        )

    maximum_percent_error = 0.0
    for spec in specs:
        if spec.role != "derived_baseline_percent_change":
            continue
        assert spec.baseline_column is not None and spec.endpoint_column is not None
        baseline = frame[spec.baseline_column].to_numpy(dtype=float)
        endpoint = frame[spec.endpoint_column].to_numpy(dtype=float)
        if np.any(baseline == 0):
            raise InventoryError(
                f"baseline contains zero, so {spec.column} cannot be verified safely"
            )
        expected = 100.0 * (endpoint - baseline) / baseline
        observed = frame[spec.column].to_numpy(dtype=float)
        error = float(np.max(np.abs(expected - observed)))
        maximum_percent_error = max(maximum_percent_error, error)
    if maximum_percent_error > 1e-9:
        raise InventoryError(
            "materialized percent-change columns disagree with absolute measurements: "
            f"maximum_abs_error={maximum_percent_error:.12g}"
        )

    used_range, formula_cells = _read_excel_metadata(path, sheet)
    if formula_cells:
        raise InventoryError(
            f"locked radiomics workbook unexpectedly contains {formula_cells} formula cells"
        )
    audit = WorkbookAudit(
        path=path,
        sheet=sheet,
        sha256=digest,
        rows=len(frame),
        columns=len(frame.columns),
        used_range=used_range,
        formula_cells=formula_cells,
        duplicate_patient_rows=duplicate_rows,
        duplicate_full_rows=duplicate_full_rows,
    )
    return frame, ids, audit, maximum_percent_error


def _read_clinical_workbook(
    config: Mapping[str, Any], radiomics_ids: pd.Series
) -> ClinicalAudit:
    source = config["source"]
    path = _required_path(source, "clinical_workbook")
    sheet = _required_text(source, "clinical_sheet")
    digest = _validate_sha(
        path, _required_text(source, "clinical_sha256"), "clinical workbook"
    )
    excel = pd.ExcelFile(path, engine="openpyxl")
    if excel.sheet_names != [sheet]:
        raise InventoryError(
            "clinical workbook sheet contract failed: expected exactly one configured sheet"
        )
    frame = pd.read_excel(excel, sheet_name=sheet)
    if tuple(str(column) for column in frame.columns) != EXPECTED_CLINICAL_COLUMNS:
        raise InventoryError("clinical workbook column schema does not match the locked schema")
    if frame.shape != (EXPECTED_CLINICAL_ROWS, len(EXPECTED_CLINICAL_COLUMNS)):
        raise InventoryError(
            "clinical shape mismatch: "
            f"expected {(EXPECTED_CLINICAL_ROWS, len(EXPECTED_CLINICAL_COLUMNS))}, "
            f"observed {frame.shape}"
        )
    clinical_ids = _normalise_six_digit_ids(
        frame["Patient_ID"], label="clinical workbook", expected_rows=EXPECTED_CLINICAL_ROWS
    )
    rad_set = set(radiomics_ids.tolist())
    clinical_set = set(clinical_ids.tolist())
    overlap = len(rad_set & clinical_set)
    if overlap != EXPECTED_RAD_ROWS or not rad_set.issubset(clinical_set):
        raise InventoryError(
            "clinical/radiomics exact-ID overlap failed: "
            f"expected {EXPECTED_RAD_ROWS}, observed {overlap}"
        )
    subset = frame.loc[clinical_ids.isin(rad_set)].copy()
    target_summary: list[dict[str, Any]] = []
    for target in ("pCR", "HR", "HER2", "MP"):
        values = pd.to_numeric(subset[target], errors="coerce")
        missing = int(values.isna().sum())
        invalid = int((~values.dropna().isin([0, 1])).sum())
        if missing or invalid:
            raise InventoryError(
                f"clinical target {target} failed: missing={missing}, invalid_binary={invalid}"
            )
        positive = int((values == 1).sum())
        target_summary.append(
            {
                "target": target,
                "n": len(values),
                "n_valid": int(values.notna().sum()),
                "n_missing": missing,
                "n_positive": positive,
                "positive_pct": 100.0 * positive / len(values),
            }
        )
    hr = pd.to_numeric(subset["HR"], errors="raise").astype(int)
    her2 = pd.to_numeric(subset["HER2"], errors="raise").astype(int)
    subtype = hr.map({0: "HR-", 1: "HR+"}) + "/" + her2.map(
        {0: "HER2-", 1: "HER2+"}
    )
    subtype_summary = tuple(
        {
            "subtype": label,
            "n": int((subtype == label).sum()),
            "pct": 100.0 * int((subtype == label).sum()) / len(subtype),
        }
        for label in ("HR-/HER2-", "HR-/HER2+", "HR+/HER2-", "HR+/HER2+")
    )
    return ClinicalAudit(
        path=path,
        sheet=sheet,
        sha256=digest,
        rows=len(frame),
        columns=len(frame.columns),
        radiomics_overlap=overlap,
        target_summary=tuple(target_summary),
        subtype_summary=subtype_summary,
    )


def _read_mri_coverage(
    config: Mapping[str, Any], radiomics_ids: pd.Series
) -> MriCoverage:
    source = config["source"]
    path = _required_path(source, "mri_fold_manifest")
    digest = _validate_sha(
        path,
        _required_text(source, "mri_fold_manifest_sha256"),
        "MRI fold manifest",
    )
    frame = pd.read_csv(path, usecols=["patient_id", "fold", "split"])
    if frame[["patient_id", "fold", "split"]].isna().any().any():
        raise InventoryError("MRI fold manifest contains missing contract fields")
    canonical = frame["patient_id"].astype(str).str.extract(
        r"^(?:ISPY2-|ACRIN-6698-)(\d{6})$", expand=False
    )
    invalid_rows = int(canonical.isna().sum())
    if invalid_rows:
        raise InventoryError(
            f"MRI fold manifest contains {invalid_rows} noncanonical patient IDs"
        )
    per_patient = frame.assign(_canonical=canonical).groupby("_canonical", sort=False)
    patient_rows = per_patient.size()
    if len(patient_rows) != EXPECTED_MRI_PATIENTS:
        raise InventoryError(
            f"MRI cohort count mismatch: expected {EXPECTED_MRI_PATIENTS}, observed {len(patient_rows)}"
        )
    if not patient_rows.eq(5).all():
        raise InventoryError("each MRI patient must occur exactly once per outer fold")
    fold_counts = per_patient["fold"].nunique()
    if not fold_counts.eq(5).all():
        raise InventoryError("MRI patients do not each cover all five folds")

    mri_ids = set(patient_rows.index.astype(str).tolist())
    rad_ids = radiomics_ids.astype(str)
    rad_set = set(rad_ids.tolist())
    matched_mask = rad_ids.isin(mri_ids).to_numpy(dtype=bool)
    overlap = int(matched_mask.sum())
    if overlap != EXPECTED_MRI_OVERLAP:
        raise InventoryError(
            f"MRI/radiomics overlap mismatch: expected {EXPECTED_MRI_OVERLAP}, observed {overlap}"
        )
    return MriCoverage(
        path=path,
        sha256=digest,
        cohort_patients=len(mri_ids),
        radiomics_overlap=overlap,
        workbook_only=len(rad_set - mri_ids),
        mri_only=len(mri_ids - rad_set),
        matched_workbook_mask=matched_mask,
    )


def _leakage_concern(spec: ColumnSpec) -> str:
    if spec.role == "patient_identifier":
        return (
            "Direct identifier: exact joins/splits only; never a predictor and never publish values."
        )
    family_concern = {
        "FTV": (
            "FTV is tumor-burden/segmentation-derived and can be redundant with ROI mask geometry; "
            "fit residualization within outer train."
        ),
        "SPH": (
            "SPH is derived from FTV-mask geometry and is not independent of FTV/ROI geometry."
        ),
        "LD": (
            "LD is another tumor-burden measure; later zero/floor values are valid observations, "
            "not missing values."
        ),
        "BPE": (
            "BPE is contralateral/global; a lesion-centered MRI crop may not contain its source anatomy."
        ),
    }[spec.family]
    timing = (
        "Pretreatment-only value; still fit every transform on outer train."
        if spec.first_usable_timing == "T0"
        else f"Unavailable before {spec.first_usable_timing}; using it earlier is future-visit leakage."
    )
    if spec.first_usable_timing == "T3":
        timing += " T3 is late/pre-surgery and must be labeled as such."
    if spec.role == "derived_baseline_percent_change":
        timing += (
            " Deterministic from T0 and the endpoint; do not count it as an independent measurement "
            "or duplicate a pipeline-derived change feature."
        )
    return f"{timing} {family_concern}"


def _inventory_frame(
    frame: pd.DataFrame,
    specs: Sequence[ColumnSpec],
    mri: MriCoverage,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    matched = frame.loc[mri.matched_workbook_mask]
    for spec in specs:
        series = frame[spec.column]
        matched_series = matched[spec.column]
        nonmissing = int(series.notna().sum())
        matched_nonmissing = int(matched_series.notna().sum())
        numeric = pd.to_numeric(series, errors="coerce")
        records.append(
            {
                "column": spec.column,
                "family": spec.family,
                "visit": spec.visit,
                "unit": spec.unit,
                "role": spec.role,
                "definition": spec.definition,
                "source_dtype": str(series.dtype),
                "first_usable_timing": spec.first_usable_timing,
                "independent_measurement": spec.independent_measurement,
                "n_rows": len(frame),
                "n_non_missing": nonmissing,
                "n_missing": len(frame) - nonmissing,
                "missingness_fraction": (len(frame) - nonmissing) / len(frame),
                "missingness_pct": 100.0 * (len(frame) - nonmissing) / len(frame),
                "workbook_patient_coverage_n": nonmissing,
                "workbook_patient_coverage_pct": 100.0 * nonmissing / len(frame),
                "mri_complete4_cohort_n": mri.cohort_patients,
                "mri_overlap_patient_coverage_n": matched_nonmissing,
                "mri_overlap_patient_coverage_pct": (
                    100.0 * matched_nonmissing / mri.cohort_patients
                ),
                "n_zero": int((numeric == 0).sum()) if spec.family != "ID" else 0,
                "possible_leakage_concern": _leakage_concern(spec),
            }
        )
    return pd.DataFrame.from_records(records)


def _missingness_frame(
    frame: pd.DataFrame,
    specs: Sequence[ColumnSpec],
    mri: MriCoverage,
) -> pd.DataFrame:
    matched = frame.loc[mri.matched_workbook_mask]
    records: list[dict[str, Any]] = []
    for spec in specs:
        source_valid = int(frame[spec.column].notna().sum())
        matched_valid = int(matched[spec.column].notna().sum())
        scopes = (
            (
                "source_workbook",
                len(frame),
                source_valid,
                "cell_missingness",
                "One source row per workbook patient.",
            ),
            (
                "complete4_mri_cohort",
                mri.cohort_patients,
                matched_valid,
                "structural_source_unavailability",
                "Exact six-digit-ID overlap with the complete-four-visit MRI cohort.",
            ),
            (
                "mri_matched_reference",
                mri.radiomics_overlap,
                matched_valid,
                "cell_missingness_within_matched_subset",
                "Only exact MRI/radiomics matched patients.",
            ),
        )
        for scope, total, valid, missingness_type, basis in scopes:
            missing = total - valid
            records.append(
                {
                    "scope": scope,
                    "column": spec.column,
                    "family": spec.family,
                    "visit": spec.visit,
                    "unit": spec.unit,
                    "role": spec.role,
                    "n_patients": total,
                    "n_valid_patients": valid,
                    "n_missing_patients": missing,
                    "missingness_fraction": missing / total,
                    "missingness_pct": 100.0 * missing / total,
                    "patient_coverage_fraction": valid / total,
                    "patient_coverage_pct": 100.0 * valid / total,
                    "missingness_type": missingness_type,
                    "coverage_basis": basis,
                    "possible_leakage_concern": _leakage_concern(spec),
                }
            )
    return pd.DataFrame.from_records(records)


def _markdown_escape(value: Any) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")


def _markdown_table(headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> str:
    rendered = ["| " + " | ".join(_markdown_escape(v) for v in headers) + " |"]
    rendered.append("|" + "|".join("---" for _ in headers) + "|")
    for row in rows:
        rendered.append("| " + " | ".join(_markdown_escape(v) for v in row) + " |")
    return "\n".join(rendered)


def _render_report(
    audit: WorkbookAudit,
    clinical: ClinicalAudit,
    mri: MriCoverage,
    inventory: pd.DataFrame,
    maximum_percent_error: float,
) -> str:
    inventory_rows = []
    for row in inventory.itertuples(index=False):
        inventory_rows.append(
            (
                f"`{row.column}`",
                row.family,
                row.visit,
                row.unit,
                row.role,
                f"{row.n_missing}/{row.n_rows} ({row.missingness_pct:.2f}%)",
                f"{row.workbook_patient_coverage_n}/{audit.rows}",
                f"{row.mri_overlap_patient_coverage_n}/{mri.cohort_patients}",
                row.possible_leakage_concern,
            )
        )
    target_rows = [
        (
            row["target"],
            row["n_valid"],
            row["n_missing"],
            row["n_positive"],
            f"{row['positive_pct']:.2f}%",
        )
        for row in clinical.target_summary
    ]
    subtype_rows = [
        (row["subtype"], row["n"], f"{row['pct']:.2f}%")
        for row in clinical.subtype_summary
    ]

    return f"""# Classical DCE radiomics feature inventory

本报告由 `scripts/inventory.py` 从锁定的只读输入确定性生成。报告和两个 CSV 只含 schema 与 aggregate count，不含任何患者 ID 或 patient-level measurement。

## 1. 权威输入与 fail-closed 核验

| 输入 | 路径 | SHA-256 | 结构 |
|---|---|---|---|
| DCE measurement workbook | `{audit.path}` | `{audit.sha256}` | `{audit.sheet}`；{audit.rows}×{audit.columns}；used range `{audit.used_range}` |
| Clinical workbook | `{clinical.path}` | `{clinical.sha256}` | `{clinical.sheet}`；{clinical.rows}×{clinical.columns} |
| Complete-four-visit MRI fold manifest | `{mri.path}` | `{mri.sha256}` | {mri.cohort_patients} unique patients；5 outer folds |

Radiomics workbook 只有一个 visible sheet；公式单元格 {audit.formula_cells}，重复患者 {audit.duplicate_patient_rows}，重复整行 {audit.duplicate_full_rows}。`{ID_COLUMN}` 为 {audit.rows} 个唯一、非缺失、恰好六位的数值 ID。匹配仅允许 workbook 六位 ID 与 canonical MRI ID 后缀精确等值，不做 fuzzy matching，也不在任何输出中物化 ID。

12 个 `*_pch_T0_Tk` 列均通过 `100 × (X_Tk − X_T0) / X_T0` 复算；全表最大绝对误差 `{maximum_percent_error:.3e}`。因此它们是派生字段，不是 12 个额外独立 measurement。

## 2. 实际 feature family

- **F / FTV**：functional tumor volume，单位 cc。
- **D / LD**：longest diameter；工作簿没有声明绝对单位。
- **S / SPH**：由 3-D FTV mask 得到的 sphericity，无量纲。
- **B / BPE**：对侧乳腺中央五层纤维腺体 mean early PE；工作簿没有独立声明绝对 scale/unit。
- **Other**：无。源表不是高维 PyRadiomics texture export。

绝对 measurement 共 16 列（4 family × T0–T3）；另有 12 列相对 T0 的 materialized percent change。`NONFTV = D + S + B`，`FULL = F + NONFTV`。

## 3. 逐列 inventory

{_markdown_table(
        ("Column", "Family", "Visit", "Unit", "Role", "Source missing", "Workbook coverage", "MRI-cohort coverage", "Possible leakage concern"),
        inventory_rows,
    )}

完整机器可读版本：`features/radiomics_feature_inventory.csv`。

## 4. Missingness 与 matched population

- 源 workbook：{audit.rows}/{audit.rows} patients 的 29 列全部非缺失；这是 workbook 内 complete-case，不代表完整 MRI cohort 无缺失。
- Clinical：全部 {audit.rows} workbook patients 可精确匹配 clinical row；pCR、HR、HER2、MP 均 0 missing。
- MRI reference：{mri.radiomics_overlap}/{mri.cohort_patients} ({100.0 * mri.radiomics_overlap / mri.cohort_patients:.2f}%) 有 radiomics；{mri.mri_only}/{mri.cohort_patients} ({100.0 * mri.mri_only / mri.cohort_patients:.2f}%) 对全部 28 measurement 结构性缺失。另有 {mri.workbook_only} 个 workbook patients 不属于 complete-four-visit MRI cohort。
- LD 零值数 T0/T1/T2/T3 为 {int(inventory.loc[inventory.column.eq('LD_T0'), 'n_zero'].iloc[0])}/{int(inventory.loc[inventory.column.eq('LD_T1'), 'n_zero'].iloc[0])}/{int(inventory.loc[inventory.column.eq('LD_T2'), 'n_zero'].iloc[0])}/{int(inventory.loc[inventory.column.eq('LD_T3'), 'n_zero'].iloc[0])}；BPE 为 {int(inventory.loc[inventory.column.eq('BPE_5slice_mean_T0'), 'n_zero'].iloc[0])}/{int(inventory.loc[inventory.column.eq('BPE_5slice_mean_T1'), 'n_zero'].iloc[0])}/{int(inventory.loc[inventory.column.eq('BPE_5slice_mean_T2'), 'n_zero'].iloc[0])}/{int(inventory.loc[inventory.column.eq('BPE_5slice_mean_T3'), 'n_zero'].iloc[0])}。这些是观察值，不得改写为 NA。

逐 scope、逐 column 的机器可读统计见 `metrics/missingness.csv`。其中 `complete4_mri_cohort` 明确把没有 source row 的患者记为 structural source unavailability，而不是在 384-row workbook 内伪造 NA。

## 5. Aggregate target coverage（384 primary population）

{_markdown_table(("Target", "Valid", "Missing", "Positive", "Positive rate"), target_rows)}

{_markdown_table(("HR/HER2 subtype", "N", "Rate"), subtype_rows)}

Radiomics workbook 本身不含 outcome、molecular subtype、treatment 或 demographic 字段；这些 target aggregate 仅来自 SHA-locked clinical workbook 的 exact-ID intersection。

## 6. Timing-safe 使用边界

- T0 只能用 T0；T1 只能用 T0/T1；T2 只能用 T0–T2；T3 才能用 T0–T3，并必须标记为 `late/pre-surgery`。
- `*_pch_T0_Tk` 只有在 Tk 已观察后才可使用，并与其两个 endpoint 确定性重复。正式 pipeline 应从当时可见的 absolute value 按预注册公式重建 change。
- winsorization、log transform、scaling、imputation、feature selection 与 FTV residualization 全部只在 outer train 拟合。
- FTV 与 ROI/mask geometry 高度同源；SPH 直接依赖 FTV-mask geometry；LD 也是 burden proxy。不能把它们的增益直接解释成独立 biological signal。
- BPE 来自对侧/全乳背景组织。与 lesion-centered LOCAL latent 比较时必须注明 anatomy/input mismatch。
- 含 clinical label 的派生表不得使用“排除 label 的黑名单”；建模代码必须使用显式 predictor allowlist。

## 7. Population recommendation

Primary classical C/F/N/FULL、residualization、family ablation 与 HR/HER2 probe 使用全部 {audit.rows} 人；所有 paired model 在同一 complete population 上比较，结论不依赖 current MRI availability。MRI latent reference 与其 matched sensitivity analysis 使用严格 {mri.radiomics_overlap} 人 subset，并复用 MRI folds。二者必须分别标记为 `primary_384` 与 `mri_matched_375`，不得混报 AUROC。
"""


def _atomic_write_csv(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        newline="",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary = Path(handle.name)
        frame.to_csv(handle, index=False, float_format="%.12g")
    try:
        os.replace(temporary, path)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def _atomic_write_text(text: str, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        newline="\n",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temporary = Path(handle.name)
        handle.write(text.rstrip() + "\n")
    try:
        os.replace(temporary, path)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def run_inventory(config_path: Path, output_root: Path) -> InventoryResult:
    """Validate all inputs and write aggregate inventory outputs atomically."""

    config = _load_config(config_path)
    specs = _build_column_specs(config)
    frame, radiomics_ids, audit, maximum_percent_error = _read_radiomics_workbook(
        config, specs
    )
    clinical = _read_clinical_workbook(config, radiomics_ids)
    mri = _read_mri_coverage(config, radiomics_ids)
    inventory = _inventory_frame(frame, specs, mri)
    missingness = _missingness_frame(frame, specs, mri)
    report = _render_report(
        audit, clinical, mri, inventory, maximum_percent_error
    )

    inventory_path = output_root / "features" / "radiomics_feature_inventory.csv"
    missingness_path = output_root / "metrics" / "missingness.csv"
    report_path = output_root / "reports" / "radiomics_feature_inventory.md"
    _atomic_write_csv(inventory, inventory_path)
    _atomic_write_csv(missingness, missingness_path)
    _atomic_write_text(report, report_path)
    return InventoryResult(
        inventory_path=inventory_path,
        missingness_path=missingness_path,
        report_path=report_path,
        inventory_rows=len(inventory),
        missingness_rows=len(missingness),
        workbook_patients=audit.rows,
        mri_overlap=mri.radiomics_overlap,
    )


def _parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG,
        help=f"Experiment JSON (default: {DEFAULT_CONFIG})",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=EXPERIMENT_ROOT,
        help=f"Experiment output root (default: {EXPERIMENT_ROOT})",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        result = run_inventory(args.config.resolve(), args.output_root.resolve())
    except (InventoryError, OSError, ValueError) as exc:
        print(f"INVENTORY_VALIDATION_FAILED: {exc}", file=sys.stderr)
        return 2
    summary = {
        "status": "ok",
        "workbook_patients": result.workbook_patients,
        "mri_overlap": result.mri_overlap,
        "inventory_rows": result.inventory_rows,
        "missingness_rows": result.missingness_rows,
        "inventory_path": str(result.inventory_path),
        "missingness_path": str(result.missingness_path),
        "report_path": str(result.report_path),
        "contains_patient_level_values": False,
    }
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

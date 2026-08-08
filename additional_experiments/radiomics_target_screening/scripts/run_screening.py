#!/usr/bin/env python3
"""Outcome-free I-SPY2 radiomics target screening.

本脚本只读取 measurement workbook、锁定的 patient split 和既有只读审计结果；
不会读取 pCR/clinical outcome，不会训练或修改任何神经网络。
"""

from __future__ import annotations

import hashlib
import json
import math
import re
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from scipy.stats import pearsonr, spearmanr
from sklearn.linear_model import Ridge
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler


REPO_ROOT = Path(__file__).resolve().parents[3]
EXP_ROOT = Path(__file__).resolve().parents[1]
METRICS = EXP_ROOT / "metrics"
FIGURES = EXP_ROOT / "figures"
REPORTS = EXP_ROOT / "reports"

WORKBOOK = Path("/data/data/Breast_Cancer/I-SPY2/Multi-feature-MRI-NACT-Data.xlsx")
EXPECTED_WORKBOOK_SHA256 = "f714c7784b1e57daa74d7cfb20db71cd432b4e4596b9b4eacdd5a76b7f8a58dc"
SHEET = "datawith4visits"
FOLD_MANIFEST = Path(
    "/data/data/Preprocessed/I-SPY2/"
    "_matched_breastdcedl_t0_dicomrepair_rgb224_seed2026/"
    "matched_patient_cv_splits_seed2026.csv"
)
EXPECTED_FOLD_SHA256 = "143e482d711225c0611006d99bd7345d2fa1a5c16c65fbaf8399341a0d26aa38"
OVERLAP = REPO_ROOT / "additional_experiments/radiomics_next_change/data_audit/radiomics_patient_overlap.csv"
EXPECTED_OVERLAP_SHA256 = "91b575c9e7e351312b8181a091bdffd2d1f61b88b5a98ac3d78d54c94b63da6b"
OSRA_METRICS = (
    REPO_ROOT
    / "additional_experiments/observed_state_radiomics_audit/metrics/final_analysis/oof_metrics.csv"
)
DGRS_PROBES = (
    REPO_ROOT
    / "additional_experiments/direct_grounded_response_state/metrics/final/probe_oof_metrics.csv"
)

VISITS = ("T0", "T1", "T2", "T3")
TRANSITIONS = ("T0→T1", "T1→T2", "T2→T3")
CANDIDATES = ("LD", "SPH", "BPE")
ALL_FEATURES = ("FTV", *CANDIDATES)

FEATURES: OrderedDict[str, dict[str, Any]] = OrderedDict(
    {
        "FTV": {
            "columns": (
                "VOLUME_TUM_BLU_V10",
                "VOLUME_TUM_BLU_V20",
                "VOLUME_TUM_BLU_V30",
                "VOLUME_TUM_BLU_V40",
            ),
            "role": "reference_target",
            "unit": "cc（随附 DICOM 说明；工作簿未单列单位）",
            "definition": "满足 PE/SER 阈值的增强组织功能性肿瘤体积",
            "region": "病灶 VOI/FTV 分析区域",
        },
        "LD": {
            "columns": ("LD_T0", "LD_T1", "LD_T2", "LD_T3"),
            "role": "formal_candidate",
            "unit": "工作簿未明示",
            "definition": "影像报告中的肿瘤最长径（longest diameter）",
            "region": "病灶本体",
        },
        "SPH": {
            "columns": ("SPHERICITY_T0", "SPHERICITY_T1", "SPHERICITY_T2", "SPHERICITY_T3"),
            "role": "formal_candidate",
            "unit": "无量纲，范围理论上 0–1",
            "definition": "等体积球表面积 / 3D FTV tumor mask 表面积",
            "region": "病灶 3D FTV mask/边界",
        },
        "BPE": {
            "columns": (
                "BPE_5slice_mean_T0",
                "BPE_5slice_mean_T1",
                "BPE_5slice_mean_T2",
                "BPE_5slice_mean_T3",
            ),
            "role": "formal_candidate",
            "unit": "percent enhancement；工作簿未单列单位",
            "definition": "对侧乳腺中央连续 5 层纤维腺体组织的平均早期 PE",
            "region": "对侧乳腺中央纤维腺体组织",
        },
    }
)

PCH_COLUMNS: OrderedDict[str, tuple[str, ...]] = OrderedDict(
    {
        "FTV": ("FTV_pch_T0_T1", "FTV_pch_T0_T2", "FTV_pch_T0_T3"),
        "SPH": (
            "Sphericity_pch_T0_T1",
            "Sphericity_pch_T0_T2",
            "Sphericity_pch_T0_T3",
        ),
        "LD": ("LD_pch_T0_T1", "LD_pch_T0_T2", "LD_pch_T0_T3"),
        "BPE": ("BPE_pch_T0_T1", "BPE_pch_T0_T2", "BPE_pch_T0_T3"),
    }
)

EXPECTED_COLUMNS = (
    "CLINICAL-TRIAL-SUBJECT-ID",
    *FEATURES["FTV"]["columns"],
    *FEATURES["SPH"]["columns"],
    *FEATURES["LD"]["columns"],
    *FEATURES["BPE"]["columns"],
    *PCH_COLUMNS["FTV"],
    *PCH_COLUMNS["SPH"],
    *PCH_COLUMNS["LD"],
    *PCH_COLUMNS["BPE"],
)

PRIMARY_PAPER = "https://pmc.ncbi.nlm.nih.gov/articles/PMC7695723/"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def finite_pair(x: Iterable[float], y: Iterable[float]) -> tuple[np.ndarray, np.ndarray]:
    xa = np.asarray(list(x), dtype=np.float64)
    ya = np.asarray(list(y), dtype=np.float64)
    valid = np.isfinite(xa) & np.isfinite(ya)
    return xa[valid], ya[valid]


def correlations(x: Iterable[float], y: Iterable[float]) -> tuple[int, float, float]:
    xa, ya = finite_pair(x, y)
    if xa.size < 3 or np.unique(xa).size < 2 or np.unique(ya).size < 2:
        return int(xa.size), math.nan, math.nan
    return int(xa.size), float(spearmanr(xa, ya).statistic), float(pearsonr(xa, ya).statistic)


def iqr(values: Iterable[float]) -> float:
    array = np.asarray(list(values), dtype=np.float64)
    array = array[np.isfinite(array)]
    if not array.size:
        return math.nan
    q1, q3 = np.quantile(array, (0.25, 0.75))
    return float(q3 - q1)


def fmt(value: Any, digits: int = 3) -> str:
    if value is None or (isinstance(value, (float, np.floating)) and not np.isfinite(value)):
        return "NA"
    if isinstance(value, (float, np.floating)):
        return f"{float(value):.{digits}f}"
    return str(value)


def markdown_table(frame: pd.DataFrame) -> str:
    if frame.empty:
        return "（无记录）"
    clean = frame.copy()
    clean.columns = [str(column).replace("|", "\\|") for column in clean.columns]
    rows = []
    rows.append("| " + " | ".join(clean.columns) + " |")
    rows.append("|" + "|".join("---" for _ in clean.columns) + "|")
    for row in clean.itertuples(index=False, name=None):
        cells = []
        for value in row:
            if pd.isna(value):
                text = "NA"
            elif isinstance(value, (float, np.floating)):
                text = fmt(value)
            else:
                text = str(value)
            cells.append(text.replace("|", "\\|").replace("\n", " "))
        rows.append("| " + " | ".join(cells) + " |")
    return "\n".join(rows)


def save_csv(frame: pd.DataFrame, name: str) -> None:
    frame.to_csv(METRICS / name, index=False)


def read_and_validate_workbook() -> tuple[pd.DataFrame, list[str], dict[str, Any]]:
    if not WORKBOOK.is_file():
        raise FileNotFoundError(WORKBOOK)
    if sha256(WORKBOOK) != EXPECTED_WORKBOOK_SHA256:
        raise ValueError("measurement workbook SHA256 与锁定输入不一致")
    excel = pd.ExcelFile(WORKBOOK)
    sheet_names = list(excel.sheet_names)
    if sheet_names != [SHEET]:
        raise ValueError(f"工作簿 sheet 漂移: {sheet_names}")
    frame = pd.read_excel(WORKBOOK, sheet_name=SHEET)
    if tuple(frame.columns) != EXPECTED_COLUMNS:
        raise ValueError("工作簿列结构与真实检查时不一致")
    if frame.shape != (384, 29):
        raise ValueError(f"工作簿维度漂移: {frame.shape}")
    if frame.isna().any().any():
        raise ValueError("complete-4-visit 工作簿出现意外缺失")
    if frame.duplicated().any():
        raise ValueError("工作簿出现重复整行")
    ids = frame["CLINICAL-TRIAL-SUBJECT-ID"]
    id_strings = ids.astype(str).str.strip().str.zfill(6)
    if not id_strings.str.fullmatch(r"\d{6}").all() or id_strings.duplicated().any():
        raise ValueError("患者 ID 非唯一六位整数")
    for column in frame.columns[1:]:
        values = pd.to_numeric(frame[column], errors="coerce").to_numpy(dtype=np.float64)
        if not np.isfinite(values).all():
            raise ValueError(f"列含 non-finite: {column}")

    pch_max_errors: dict[str, float] = {}
    for feature, pch_columns in PCH_COLUMNS.items():
        raw_columns = FEATURES[feature]["columns"]
        baseline = frame[raw_columns[0]].to_numpy(dtype=np.float64)
        if np.any(baseline == 0):
            raise ValueError(f"{feature} baseline 含零，无法验证 pch")
        errors = []
        for end_index, pch_column in enumerate(pch_columns, start=1):
            expected = 100.0 * (frame[raw_columns[end_index]].to_numpy(dtype=np.float64) - baseline) / baseline
            observed = frame[pch_column].to_numpy(dtype=np.float64)
            errors.append(float(np.max(np.abs(expected - observed))))
        pch_max_errors[feature] = max(errors)
    if max(pch_max_errors.values()) > 1e-8:
        raise ValueError(f"pch 数值关系验证失败: {pch_max_errors}")

    workbook = load_workbook(WORKBOOK, read_only=False, data_only=False)
    worksheet = workbook[SHEET]
    metadata = {
        "sheet_names": sheet_names,
        "sheet_state": worksheet.sheet_state,
        "dimension": worksheet.calculate_dimension(),
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "merged_ranges": len(worksheet.merged_cells.ranges),
        "formula_cells": sum(
            1 for row in worksheet.iter_rows() for cell in row if cell.data_type == "f"
        ),
        "duplicate_rows": int(frame.duplicated().sum()),
        "duplicate_rows_without_id": int(frame.drop(columns=[frame.columns[0]]).duplicated().sum()),
        "duplicate_patient_ids": int(id_strings.duplicated().sum()),
        "pch_max_abs_error": pch_max_errors,
    }
    workbook.close()
    return frame, sheet_names, metadata


def feature_and_visit_for_column(column: str) -> tuple[str, str, str]:
    if column == "CLINICAL-TRIAL-SUBJECT-ID":
        return "patient_identifier", "", ""
    for feature, spec in FEATURES.items():
        if column in spec["columns"]:
            visit = VISITS[spec["columns"].index(column)]
            return spec["role"], feature, visit
    for feature, columns in PCH_COLUMNS.items():
        if column in columns:
            end_visit = VISITS[columns.index(column) + 1]
            return "derived_baseline_percent_change_non_candidate", feature, f"T0→{end_visit}"
    raise KeyError(column)


def build_schema(frame: pd.DataFrame, metadata: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for column in frame.columns:
        role, feature, visit = feature_and_visit_for_column(column)
        series = frame[column]
        numeric = pd.to_numeric(series, errors="coerce").to_numpy(dtype=np.float64)
        finite = numeric[np.isfinite(numeric)]
        unit = "不适用"
        definition = "患者唯一六位 ClinicalTrialSubjectID"
        if feature:
            unit = "%（由 raw measurement 相对 T0 派生）" if role.startswith("derived") else FEATURES[feature]["unit"]
            definition = (
                "100 × (X_Tk − X_T0) / X_T0；非独立 measurement"
                if role.startswith("derived")
                else FEATURES[feature]["definition"]
            )
        rows.append(
            {
                "file": str(WORKBOOK),
                "file_sha256": sha256(WORKBOOK),
                "sheet": SHEET,
                "sheet_dimension": metadata["dimension"],
                "column": column,
                "dtype": str(series.dtype),
                "role": role,
                "feature": feature,
                "visit_or_interval": visit,
                "unit": unit,
                "definition": definition,
                "n_rows": len(series),
                "n_nonmissing": int(series.notna().sum()),
                "n_missing": int(series.isna().sum()),
                "missing_pct": float(100.0 * series.isna().mean()),
                "n_infinite": int(np.isinf(numeric).sum()),
                "n_unique": int(series.nunique(dropna=True)),
                "n_zero": int(np.sum(finite == 0)),
                "min": float(np.min(finite)),
                "median": float(np.median(finite)),
                "max": float(np.max(finite)),
            }
        )
    return pd.DataFrame(rows)


def measurement_table(frame: pd.DataFrame) -> pd.DataFrame:
    output = pd.DataFrame(
        {
            "trial_id": frame["CLINICAL-TRIAL-SUBJECT-ID"].astype(str).str.strip().str.zfill(6)
        }
    )
    for feature, spec in FEATURES.items():
        for visit, column in zip(VISITS, spec["columns"]):
            output[f"{feature}_{visit}"] = pd.to_numeric(frame[column], errors="raise").astype(float)
    return output


def strict_cohort_mapping(measurements: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    if sha256(FOLD_MANIFEST) != EXPECTED_FOLD_SHA256:
        raise ValueError("seed-2026 fold manifest SHA256 不匹配")

    # 明确只读 patient_id/fold/split；不读取 label_pcr。
    folds = pd.read_csv(
        FOLD_MANIFEST,
        usecols=["patient_id", "fold", "split"],
        dtype={"patient_id": str, "fold": int, "split": str},
    )
    if len(folds) != 808 * 5 or set(folds["fold"]) != set(range(5)):
        raise ValueError("fold manifest 不是锁定的 808×5 long form")
    if not set(folds["split"]).issubset({"train", "val", "test"}):
        raise ValueError("fold manifest 含未知 split")
    for fold in range(5):
        subset = folds[folds["fold"] == fold]
        if len(subset) != 808 or subset["patient_id"].duplicated().any():
            raise ValueError(f"fold {fold} 未恰好覆盖 808 名唯一患者")
    test_counts = folds.assign(is_test=folds["split"].eq("test")).groupby("patient_id")["is_test"].sum()
    if not test_counts.eq(1).all():
        raise ValueError("每位患者不是恰好一次 outer test")

    # 使用既有已验证 mapping，但先锁定文件，再重新严格验证 canonical ID 后缀与 clinical ID 等值。
    if sha256(OVERLAP) != EXPECTED_OVERLAP_SHA256:
        raise ValueError("既有 strict overlap mapping SHA256 不匹配")
    overlap = pd.read_csv(
        OVERLAP,
        usecols=["patient_id", "clinical_patient_id", "has_radiomics", "match_rule"],
        dtype={"patient_id": str, "clinical_patient_id": str, "match_rule": str},
    )
    if len(overlap) != 808 or overlap["patient_id"].duplicated().any():
        raise ValueError("既有 overlap mapping 未恰好覆盖 808 名唯一患者")
    pattern = re.compile(r"^(?:ISPY2-|ACRIN-6698-)(\d{6})$")
    suffixes = []
    for patient_id in overlap["patient_id"]:
        match = pattern.fullmatch(str(patient_id).strip())
        if match is None:
            raise ValueError(f"canonical patient_id 非严格格式: {patient_id}")
        suffixes.append(match.group(1))
    overlap["trial_id"] = suffixes
    clinical = overlap["clinical_patient_id"].astype(str).str.replace(r"\.0$", "", regex=True).str.zfill(6)
    if not clinical.str.fullmatch(r"\d{6}").all() or not np.array_equal(clinical, overlap["trial_id"]):
        raise ValueError("patient_id 后缀与 clinical_patient_id 不一致")
    if overlap["trial_id"].duplicated().any():
        raise ValueError("808 MRI cohort 的六位 trial ID 不唯一")
    has_radiomics = overlap["has_radiomics"].astype(str).str.strip().str.lower().map(
        {"true": True, "false": False}
    )
    if has_radiomics.isna().any():
        raise ValueError("has_radiomics 含非严格 true/false 值")
    if set(overlap["patient_id"]) != set(folds["patient_id"]):
        raise ValueError("overlap mapping patient set 与 fold manifest 不一致")
    expected_has_radiomics = overlap["trial_id"].isin(set(measurements["trial_id"]))
    if not np.array_equal(has_radiomics.to_numpy(dtype=bool), expected_has_radiomics.to_numpy(dtype=bool)):
        raise ValueError("has_radiomics 与 workbook 六位 ID 精确交集不一致")
    expected_rules = np.where(
        has_radiomics.to_numpy(dtype=bool),
        "显式六位 ClinicalTrialSubjectID 等值匹配",
        "无匹配；未做模糊匹配",
    )
    if not np.array_equal(overlap["match_rule"].astype(str).to_numpy(), expected_rules):
        raise ValueError("strict overlap mapping 的 match_rule 漂移")
    mapping = overlap.loc[has_radiomics, ["patient_id", "trial_id", "match_rule"]].copy()
    formal = measurements.merge(mapping, on="trial_id", how="inner", validate="one_to_one")
    if len(formal) != 375 or formal["patient_id"].duplicated().any():
        raise ValueError(f"严格 MRI overlap 不是 375: {len(formal)}")
    if not set(formal["patient_id"]).issubset(set(folds["patient_id"])):
        raise ValueError("measurement overlap 含 fold manifest 外患者")

    workbook_ids = set(measurements["trial_id"])
    cohort_ids = set(overlap["trial_id"])
    overlap_ids = set(formal["trial_id"])
    mapping_summary = {
        "workbook_patient_count": len(workbook_ids),
        "mri_cohort_patient_count": len(cohort_ids),
        "strict_overlap_count": len(overlap_ids),
        "workbook_only_count": len(workbook_ids - cohort_ids),
        "mri_only_count": len(cohort_ids - workbook_ids),
        "workbook_only_ids": sorted(workbook_ids - cohort_ids),
        "mapping_rule": "canonical patient_id regex 后缀 == clinical_patient_id == workbook 六位 ID；精确等值；无 fuzzy matching",
        "fold_train_overlap_counts": {
            str(fold): int(
                formal["patient_id"].isin(
                    folds.loc[(folds["fold"] == fold) & (folds["split"] == "train"), "patient_id"]
                ).sum()
            )
            for fold in range(5)
        },
    }
    return formal, folds, mapping_summary


def build_coverage(
    measurements: pd.DataFrame, formal: pd.DataFrame, mapping_summary: dict[str, Any]
) -> pd.DataFrame:
    rows = []
    for feature in ALL_FEATURES:
        columns = [f"{feature}_{visit}" for visit in VISITS]
        row: dict[str, Any] = {
            "feature": feature,
            "role": FEATURES[feature]["role"],
            "formal_candidate": feature in CANDIDATES,
            "definition": FEATURES[feature]["definition"],
            "unit": FEATURES[feature]["unit"],
            "workbook_total_patients": len(measurements),
            "mri_cohort_total_patients": mapping_summary["mri_cohort_patient_count"],
            "strict_mri_overlap_patients": len(formal),
            "workbook_only_patients": mapping_summary["workbook_only_count"],
            "mri_only_patients": mapping_summary["mri_only_count"],
        }
        for visit, column in zip(VISITS, columns):
            row[f"{visit}_available_workbook"] = int(measurements[column].notna().sum())
            row[f"{visit}_available_overlap"] = int(formal[column].notna().sum())
        row["complete_4visit_workbook"] = int(measurements[columns].notna().all(axis=1).sum())
        row["complete_4visit_overlap"] = int(formal[columns].notna().all(axis=1).sum())
        row["missing_pct_workbook"] = float(100.0 * measurements[columns].isna().to_numpy().mean())
        row["missing_pct_overlap"] = float(100.0 * formal[columns].isna().to_numpy().mean())
        for index, transition in enumerate(TRANSITIONS):
            paired = measurements[[columns[index], columns[index + 1]]].notna().all(axis=1)
            overlap_paired = formal[[columns[index], columns[index + 1]]].notna().all(axis=1)
            safe = transition.replace("→", "_to_")
            row[f"{safe}_paired_workbook"] = int(paired.sum())
            row[f"{safe}_paired_overlap"] = int(overlap_paired.sum())
        rows.append(row)
    return pd.DataFrame(rows)


def subset_for_fold(formal: pd.DataFrame, folds: pd.DataFrame, fold: int) -> pd.DataFrame:
    train_ids = set(
        folds.loc[(folds["fold"] == fold) & (folds["split"] == "train"), "patient_id"].astype(str)
    )
    return formal[formal["patient_id"].isin(train_ids)].copy()


def redundancy_tables(
    measurements: pd.DataFrame, formal: pd.DataFrame, folds: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    static_rows: list[dict[str, Any]] = []
    delta_rows: list[dict[str, Any]] = []
    scopes: list[tuple[int, str, bool, pd.DataFrame]] = [
        (-1, "all_workbook_descriptive", False, measurements)
    ]
    scopes.extend((fold, "fold_train", True, subset_for_fold(formal, folds, fold)) for fold in range(5))
    for fold, scope, selection_scope, subset in scopes:
        for candidate in CANDIDATES:
            candidate_static = []
            for visit in VISITS:
                n, rho, pearson = correlations(subset[f"{candidate}_{visit}"], subset[f"FTV_{visit}"])
                candidate_static.append(abs(rho))
                static_rows.append(
                    {
                        "scope": scope,
                        "selection_scope": selection_scope,
                        "fold": fold,
                        "candidate": candidate,
                        "visit": visit,
                        "n": n,
                        "spearman": rho,
                        "abs_spearman": abs(rho),
                        "pearson": pearson,
                        "abs_pearson": abs(pearson),
                    }
                )
            for row in static_rows[-len(VISITS) :]:
                row["fold_median_abs_spearman"] = float(np.median(candidate_static))
            candidate_delta = []
            for index, transition in enumerate(TRANSITIONS):
                delta_x = subset[f"{candidate}_{VISITS[index + 1]}"] - subset[f"{candidate}_{VISITS[index]}"]
                delta_ftv = subset[f"FTV_{VISITS[index + 1]}"] - subset[f"FTV_{VISITS[index]}"]
                n, rho, pearson = correlations(delta_x, delta_ftv)
                candidate_delta.append(abs(rho))
                delta_rows.append(
                    {
                        "scope": scope,
                        "selection_scope": selection_scope,
                        "fold": fold,
                        "candidate": candidate,
                        "transition": transition,
                        "delta_definition": "raw endpoint difference: X_end - X_start",
                        "n": n,
                        "spearman": rho,
                        "abs_spearman": abs(rho),
                        "pearson": pearson,
                        "abs_pearson": abs(pearson),
                    }
                )
            for row in delta_rows[-len(TRANSITIONS) :]:
                row["fold_median_abs_spearman"] = float(np.median(candidate_delta))
    return pd.DataFrame(static_rows), pd.DataFrame(delta_rows)


def residual_information(formal: pd.DataFrame, folds: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for fold in range(5):
        subset = subset_for_fold(formal, folds, fold)
        for candidate in CANDIDATES:
            for visit in VISITS:
                x_raw, y_raw = finite_pair(subset[f"FTV_{visit}"], subset[f"{candidate}_{visit}"])
                x = np.log1p(x_raw)
                y = np.log1p(y_raw) if candidate == "LD" else y_raw
                model = make_pipeline(StandardScaler(), Ridge(alpha=1.0))
                model.fit(x[:, None], y)
                residual = y - model.predict(x[:, None])
                y_variance = float(np.var(y, ddof=0))
                y_iqr = iqr(y)
                rows.append(
                    {
                        "task": "static",
                        "fold": fold,
                        "candidate": candidate,
                        "cell": visit,
                        "n": len(y),
                        "predictor_transform": "log1p(FTV)",
                        "target_transform": "log1p(LD)" if candidate == "LD" else "identity",
                        "model": "StandardScaler + Ridge(alpha=1.0)",
                        "fit_and_evaluation_scope": "same outer-fold training patients; descriptive explained relation",
                        "original_variance": y_variance,
                        "residual_variance": float(np.var(residual, ddof=0)),
                        "residual_variance_ratio": float(np.var(residual, ddof=0) / y_variance),
                        "original_iqr": y_iqr,
                        "residual_iqr": iqr(residual),
                        "residual_iqr_ratio": float(iqr(residual) / y_iqr),
                    }
                )
            for index, transition in enumerate(TRANSITIONS):
                x = (
                    subset[f"FTV_{VISITS[index + 1]}"] - subset[f"FTV_{VISITS[index]}"]
                ).to_numpy(dtype=np.float64)
                y = (
                    subset[f"{candidate}_{VISITS[index + 1]}"]
                    - subset[f"{candidate}_{VISITS[index]}"]
                ).to_numpy(dtype=np.float64)
                x, y = finite_pair(x, y)
                model = make_pipeline(StandardScaler(), Ridge(alpha=1.0))
                model.fit(x[:, None], y)
                residual = y - model.predict(x[:, None])
                y_variance = float(np.var(y, ddof=0))
                y_iqr = iqr(y)
                rows.append(
                    {
                        "task": "delta",
                        "fold": fold,
                        "candidate": candidate,
                        "cell": transition,
                        "n": len(y),
                        "predictor_transform": "identity raw ΔFTV",
                        "target_transform": "identity raw ΔX",
                        "model": "StandardScaler + Ridge(alpha=1.0)",
                        "fit_and_evaluation_scope": "same outer-fold training patients; descriptive explained relation",
                        "original_variance": y_variance,
                        "residual_variance": float(np.var(residual, ddof=0)),
                        "residual_variance_ratio": float(np.var(residual, ddof=0) / y_variance),
                        "original_iqr": y_iqr,
                        "residual_iqr": iqr(residual),
                        "residual_iqr_ratio": float(iqr(residual) / y_iqr),
                    }
                )
    return pd.DataFrame(rows)


def within_patient_ratio(subset: pd.DataFrame, feature: str) -> tuple[float, float, float]:
    values = subset[[f"{feature}_{visit}" for visit in VISITS]].to_numpy(dtype=np.float64)
    within = float(np.mean(np.var(values, axis=1, ddof=0)))
    total = float(np.var(values.reshape(-1), ddof=0))
    return within, total, float(within / total)


def longitudinal_variability(
    measurements: pd.DataFrame, formal: pd.DataFrame, folds: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, Any]] = []
    standardized_rows: list[dict[str, Any]] = []
    scopes: list[tuple[int, str, bool, pd.DataFrame]] = [
        (-1, "all_workbook_descriptive", False, measurements)
    ]
    scopes.extend((fold, "fold_train", True, subset_for_fold(formal, folds, fold)) for fold in range(5))
    for fold, scope, selection_scope, subset in scopes:
        for feature in ALL_FEATURES:
            within, total, ratio = within_patient_ratio(subset, feature)
            for index, transition in enumerate(TRANSITIONS):
                start = subset[f"{feature}_{VISITS[index]}"].to_numpy(dtype=np.float64)
                end = subset[f"{feature}_{VISITS[index + 1]}"].to_numpy(dtype=np.float64)
                valid = np.isfinite(start) & np.isfinite(end)
                start, end = start[valid], end[valid]
                delta = end - start
                scale = iqr(start)
                if not np.isfinite(scale) or scale <= 1e-12:
                    scale = max(float(np.std(start, ddof=0)), 1e-12)
                standardized = np.abs(delta) / scale
                near_zero_threshold = max(1e-8, 0.01 * scale)
                rows.append(
                    {
                        "scope": scope,
                        "selection_scope": selection_scope,
                        "fold": fold,
                        "feature": feature,
                        "transition": transition,
                        "n": len(delta),
                        "median_delta": float(np.median(delta)),
                        "iqr_delta": iqr(delta),
                        "median_abs_delta": float(np.median(np.abs(delta))),
                        "robust_scale_start_visit_iqr": scale,
                        "median_standardized_abs_change": float(np.median(standardized)),
                        "proportion_delta_positive": float(np.mean(delta > near_zero_threshold)),
                        "proportion_delta_negative": float(np.mean(delta < -near_zero_threshold)),
                        "proportion_near_zero": float(np.mean(np.abs(delta) <= near_zero_threshold)),
                        "near_zero_definition": (
                            "abs(Δ) <= max(1e-8, 0.01 × all-workbook/start-visit IQR)"
                            if fold == -1
                            else "abs(Δ) <= max(1e-8, 0.01 × train/start-visit IQR)"
                        ),
                        "mean_within_patient_variance": within,
                        "pooled_total_variance": total,
                        "within_to_total_variance_ratio": ratio,
                        "between_patient_variance_fraction": 1.0 - ratio,
                    }
                )
                if fold == -1:
                    standardized_rows.extend(
                        {
                            "feature": feature,
                            "transition": transition,
                            "standardized_abs_change": float(value),
                        }
                        for value in standardized
                    )
    return pd.DataFrame(rows), pd.DataFrame(standardized_rows)


def distribution_summary(measurements: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for feature in ALL_FEATURES:
        for visit in VISITS:
            values = measurements[f"{feature}_{visit}"].to_numpy(dtype=np.float64)
            values = values[np.isfinite(values)]
            p1, q1, median, q3, p99 = np.quantile(values, (0.01, 0.25, 0.5, 0.75, 0.99))
            width = float(q3 - q1)
            tolerance = max(1e-12, 0.01 * width)
            floor_fraction = float(np.mean(values <= np.min(values) + tolerance))
            ceiling_fraction = float(np.mean(values >= np.max(values) - tolerance))
            rows.append(
                {
                    "scope": "all_workbook_descriptive",
                    "feature": feature,
                    "visit": visit,
                    "n": len(values),
                    "missing_pct": 0.0,
                    "median": float(median),
                    "iqr": width,
                    "p01": float(p1),
                    "p99": float(p99),
                    "min": float(np.min(values)),
                    "max": float(np.max(values)),
                    "n_unique": int(np.unique(values).size),
                    "n_zero": int(np.sum(values == 0)),
                    "zero_fraction": float(np.mean(values == 0)),
                    "near_zero_variance_flag": bool(width <= max(1e-10, 1e-6 * max(abs(median), 1.0))),
                    "extreme_tail_ratio_p99_minus_p01_over_iqr": float((p99 - p1) / width) if width > 0 else math.inf,
                    "heavy_tail_flag": bool(width > 0 and (p99 - p1) / width > 10.0),
                    "floor_fraction": floor_fraction,
                    "floor_effect_flag": bool(floor_fraction >= 0.10),
                    "ceiling_fraction": ceiling_fraction,
                    "ceiling_effect_flag": bool(ceiling_fraction >= 0.10),
                    "outlier_policy": "未删除；仅描述和标记",
                }
            )
    return pd.DataFrame(rows)


def pairwise_matrices(measurements: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    static_rows = []
    for visit in VISITS:
        matrix = measurements[[f"{feature}_{visit}" for feature in ALL_FEATURES]].copy()
        matrix.columns = ALL_FEATURES
        correlation = matrix.corr(method="spearman")
        for row_feature in ALL_FEATURES:
            static_rows.append(
                {
                    "scope": "all_workbook_descriptive",
                    "visit": visit,
                    "feature_row": row_feature,
                    **{column: float(correlation.loc[row_feature, column]) for column in ALL_FEATURES},
                }
            )
    delta_rows = []
    for index, transition in enumerate(TRANSITIONS):
        matrix = pd.DataFrame(
            {
                feature: measurements[f"{feature}_{VISITS[index + 1]}"]
                - measurements[f"{feature}_{VISITS[index]}"]
                for feature in ALL_FEATURES
            }
        )
        correlation = matrix.corr(method="spearman")
        for row_feature in ALL_FEATURES:
            delta_rows.append(
                {
                    "scope": "all_workbook_descriptive",
                    "transition": transition,
                    "feature_row": row_feature,
                    **{column: float(correlation.loc[row_feature, column]) for column in ALL_FEATURES},
                }
            )
    return pd.DataFrame(static_rows), pd.DataFrame(delta_rows)


def shortcut_summary() -> pd.DataFrame:
    data = pd.read_csv(OSRA_METRICS)
    rows = []
    risk = {"FTV": "HIGH", "LD": "MODERATE", "SPH": "HIGH", "BPE": "LOW"}
    caveat = {
        "FTV": "9-D geometry 对 static/change 均高度可预测；reference target。",
        "LD": "static geometry 中等可预测；raw change 的历史 probe 使用 log-difference，R² 较弱。",
        "SPH": "由 3D FTV mask 表面积直接定义；static 高、后两段 change 也明显依赖 geometry。",
        "BPE": "观察到的 geometry predictability 低，但 BPE 仅有探索性点估计、无正式 bootstrap CI。",
    }
    target_map = {"FTV": "ftv", "LD": "ld", "SPH": "sphericity", "BPE": "bpe"}
    for feature in ALL_FEATURES:
        target = target_map[feature]
        static = data[
            (data["task_type"] == "static")
            & (data["model"] == "m0")
            & (data["representation"] == "mask_geometry")
            & (data["input_variant"] == "current")
            & (data["target_name"] == target)
        ].sort_values("timepoint")
        change = data[
            (data["task_type"] == "change")
            & (data["model"] == "m0")
            & (data["representation"] == "mask_geometry")
            & (data["input_variant"] == "observed_difference")
            & (data["target_name"] == target)
        ]
        rows.append(
            {
                "feature": feature,
                "shortcut_risk": risk[feature],
                "geometry_feature_definition": "9-D ROI voxel volume + normalized bbox extents/diagonal + centroid",
                "static_median_abs_spearman": float(static["spearman"].abs().median()),
                "static_median_r2": float(static["r2"].median()),
                "static_cells": "; ".join(
                    f"{row.timepoint}:rho={row.spearman:.3f},R2={row.r2:.3f}"
                    for row in static.itertuples(index=False)
                ),
                "change_target_transform_in_prior_audit": (
                    "log endpoint difference: log(end+eps)-log(start+eps)"
                    if feature in {"FTV", "LD"}
                    else "signed endpoint difference on original scale: end-start"
                ),
                "change_median_abs_spearman": float(change["spearman"].abs().median()),
                "change_median_r2": float(change["r2"].median()),
                "change_cells": "; ".join(
                    f"{row.transition}:rho={row.spearman:.3f},R2={row.r2:.3f}"
                    for row in change.itertuples(index=False)
                ),
                "formal_ci_status": (
                    "change 有正式 CI" if feature in {"FTV", "LD", "SPH"} else "仅探索点估计；无 BPE 正式 CI"
                ),
                "interpretation": caveat[feature],
                "g3_direct_mask_input": False,
                "evidence_boundary": "target geometry dependence control；不代表严格 G3 直接读取 mask",
                "evidence_source": str(OSRA_METRICS.relative_to(REPO_ROOT)),
                "evidence_source_sha256": sha256(OSRA_METRICS),
            }
        )
    return pd.DataFrame(rows)


def observability_summary() -> pd.DataFrame:
    rows = [
        {
            "feature": "FTV",
            "required_imaging_region": FEATURES["FTV"]["region"],
            "current_input_contract": "DCE7 lesion/ROI-centered [32,96,96] → encoder → GAP；无 mask channel",
            "mri_observable": "YES",
            "gate": "REFERENCE",
            "rationale": "ROI-centered DCE7 含病灶局部增强代理；既有严格 G3 probe 的 static/delta signal 支持其作为 reference。",
        },
        {
            "feature": "LD",
            "required_imaging_region": FEATURES["LD"]["region"],
            "current_input_contract": "DCE7 lesion/ROI-centered [32,96,96] → encoder → GAP；无 mask channel",
            "mri_observable": "YES_WITH_CAVEAT",
            "gate": "PASS",
            "rationale": "ROI-centered crop 含病灶定位区域且 strict-DCE7 static probe 为正，因此仅支持部分可观察；固定 crop 未逐例保证容纳完整最大径，且 ΔLD 校准很弱。",
        },
        {
            "feature": "SPH",
            "required_imaging_region": FEATURES["SPH"]["region"],
            "current_input_contract": "DCE7 lesion/ROI-centered [32,96,96] → encoder → GAP；无 mask channel",
            "mri_observable": "YES_WITH_CAVEAT",
            "gate": "PASS_WITH_HIGH_SHORTCUT_RISK",
            "rationale": "ROI-centered crop 含局部病灶边界代理且 static probe 为正，因此仅支持部分可观察；固定 crop 未逐例保证完整 3D 表面，GAP/无 mask 下 ΔSPH 几乎不可解码。",
        },
        {
            "feature": "BPE",
            "required_imaging_region": FEATURES["BPE"]["region"],
            "current_input_contract": "DCE7 lesion/ROI-centered [32,96,96] → encoder → GAP；无全乳/对侧乳腺视野保证",
            "mri_observable": "NO_WITH_CURRENT_INPUT",
            "gate": "FAIL_INPUT_MISMATCH",
            "rationale": "BPE 需要对侧乳腺中央 5 层纤维腺体；lesion-centered crop 不保证包含该区域，历史弱 signal 不能替代解剖可见性。",
        },
    ]
    source = (
        "additional_experiments/direct_grounded_response_state/reports/final_report.md; "
        "additional_experiments/observed_state_radiomics_audit/reports/final_report.md; "
        f"{PRIMARY_PAPER}"
    )
    for row in rows:
        row["measurement_definition"] = FEATURES[row["feature"]]["definition"]
        row["evidence_source"] = source
        row["assessment_type"] = "architecture feasibility gate（事实 + 保守推断）"
    return pd.DataFrame(rows)


def prior_decodability_summary() -> pd.DataFrame:
    probes = pd.read_csv(DGRS_PROBES)
    rows = []
    for model in ("G1", "G3"):
        for feature, target in (("LD", "ld"), ("SPH", "sphericity")):
            for task in ("static", "change"):
                subset = probes[
                    (probes["model"] == model)
                    & (probes["target"] == target)
                    & (probes["task"] == task)
                ]
                rows.append(
                    {
                        "feature": feature,
                        "model": model,
                        "input_contract": "strict DCE7 lesion-centered; no mask; response_state",
                        "task": task,
                        "n_cells": len(subset),
                        "macro_mean_spearman": float(subset["spearman"].mean()),
                        "macro_mean_r2": float(subset["r2"].mean()),
                        "cell_values": "; ".join(
                            f"{(row.timepoint if task == 'static' else row.transition)}:rho={row.spearman:.3f},R2={row.r2:.3f}"
                            for row in subset.itertuples(index=False)
                        ),
                        "evidence_source": str(DGRS_PROBES.relative_to(REPO_ROOT)),
                        "evidence_source_sha256": sha256(DGRS_PROBES),
                    }
                )
    rows.extend(
        [
            {
                "feature": "BPE",
                "model": "G1/G3",
                "input_contract": "strict DCE7 lesion-centered; no mask; response_state",
                "task": task,
                "n_cells": 0,
                "macro_mean_spearman": math.nan,
                "macro_mean_r2": math.nan,
                "cell_values": "UNKNOWN：DGRS 未评估 BPE；OSRA 的旧输入含 mask 且空间不匹配",
                "evidence_source": str(DGRS_PROBES.relative_to(REPO_ROOT)),
                "evidence_source_sha256": sha256(DGRS_PROBES),
            }
            for task in ("static", "change")
        ]
    )
    return pd.DataFrame(rows)


def redundancy_class(value: float) -> str:
    if value >= 0.70:
        return "HIGH"
    if value >= 0.40:
        return "MODERATE"
    return "LOW"


def fold_level_metrics(
    formal: pd.DataFrame,
    folds: pd.DataFrame,
    static: pd.DataFrame,
    delta: pd.DataFrame,
    residual: pd.DataFrame,
    variability: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    for fold in range(5):
        n_train = len(subset_for_fold(formal, folds, fold))
        for candidate in CANDIDATES:
            static_value = float(
                static[(static["fold"] == fold) & (static["candidate"] == candidate)][
                    "abs_spearman"
                ].median()
            )
            delta_value = float(
                delta[(delta["fold"] == fold) & (delta["candidate"] == candidate)][
                    "abs_spearman"
                ].median()
            )
            static_resid = residual[
                (residual["fold"] == fold)
                & (residual["candidate"] == candidate)
                & (residual["task"] == "static")
            ]
            delta_resid = residual[
                (residual["fold"] == fold)
                & (residual["candidate"] == candidate)
                & (residual["task"] == "delta")
            ]
            variation = variability[
                (variability["fold"] == fold) & (variability["feature"] == candidate)
            ]
            rows.append(
                {
                    "fold": fold,
                    "candidate": candidate,
                    "n_fold_train_measurement_patients": n_train,
                    "static_median_abs_spearman_with_ftv": static_value,
                    "static_redundancy_class": redundancy_class(static_value),
                    "delta_median_abs_spearman_with_delta_ftv": delta_value,
                    "delta_redundancy_class": redundancy_class(delta_value),
                    "static_median_residual_variance_ratio": float(
                        static_resid["residual_variance_ratio"].median()
                    ),
                    "static_median_residual_iqr_ratio": float(static_resid["residual_iqr_ratio"].median()),
                    "delta_median_residual_variance_ratio": float(
                        delta_resid["residual_variance_ratio"].median()
                    ),
                    "delta_median_residual_iqr_ratio": float(delta_resid["residual_iqr_ratio"].median()),
                    "median_standardized_abs_change": float(
                        variation["median_standardized_abs_change"].median()
                    ),
                    "within_to_total_variance_ratio": float(
                        variation["within_to_total_variance_ratio"].iloc[0]
                    ),
                    "near_zero_change_fraction_median": float(
                        variation["proportion_near_zero"].median()
                    ),
                    "formal_selection_data": "outer-fold training patients only",
                    "pcr_used": False,
                }
            )
    return pd.DataFrame(rows)


def decision_matrix(
    coverage: pd.DataFrame,
    fold_metrics: pd.DataFrame,
    shortcut: pd.DataFrame,
    observability: pd.DataFrame,
    decodability: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    overall = {
        "LD": "RECOMMENDED",
        "SPH": "POSSIBLE SECOND CHOICE",
        "BPE": "NOT RECOMMENDED WITH CURRENT INPUT",
    }
    rationales = {
        "LD": (
            "完整覆盖、当前 crop 有部分可观察证据、跨折纵向变化稳定；ΔFTV 冗余不高且 residual 较大。"
            "虽是另一种 tumor burden measure，但 shortcut 仅中等；strict-DCE7 的 ΔLD Spearman 点估计高于 ΔSPH，"
            "但二者变换口径不同且无 paired cross-target CI，不能解释为正式优效；pilot 前需做 crop-containment audit。"
        ),
        "SPH": (
            "提供 morphology 轴且 raw Δ 与 ΔFTV 最互补，但由 FTV mask 表面积直接定义，"
            "static geometry dependence 高，strict-DCE7 Δ几乎不可解码。"
        ),
        "BPE": (
            "统计上 static 最互补、residual 最大且 geometry 风险低，但 target 来自对侧乳腺中央组织，"
            "当前 lesion-centered input 解剖区域不匹配，architecture gate 淘汰。"
        ),
    }
    for candidate in CANDIDATES:
        fold_subset = fold_metrics[fold_metrics["candidate"] == candidate]
        coverage_row = coverage[coverage["feature"] == candidate].iloc[0]
        shortcut_row = shortcut[shortcut["feature"] == candidate].iloc[0]
        observable_row = observability[observability["feature"] == candidate].iloc[0]
        g3_static = decodability[
            (decodability["feature"] == candidate)
            & (decodability["model"] == "G3")
            & (decodability["task"] == "static")
        ]
        g3_delta = decodability[
            (decodability["feature"] == candidate)
            & (decodability["model"] == "G3")
            & (decodability["task"] == "change")
        ]
        rows.append(
            {
                "candidate": candidate,
                "coverage": f"{int(coverage_row['complete_4visit_overlap'])}/375 complete",
                "coverage_fraction": float(coverage_row["complete_4visit_overlap"] / 375.0),
                "static_ftv_redundancy": float(
                    fold_subset["static_median_abs_spearman_with_ftv"].median()
                ),
                "static_redundancy_class": redundancy_class(
                    float(fold_subset["static_median_abs_spearman_with_ftv"].median())
                ),
                "delta_ftv_redundancy": float(
                    fold_subset["delta_median_abs_spearman_with_delta_ftv"].median()
                ),
                "delta_redundancy_class": redundancy_class(
                    float(fold_subset["delta_median_abs_spearman_with_delta_ftv"].median())
                ),
                "static_residual_variance_ratio": float(
                    fold_subset["static_median_residual_variance_ratio"].median()
                ),
                "delta_residual_variance_ratio": float(
                    fold_subset["delta_median_residual_variance_ratio"].median()
                ),
                "longitudinal_standardized_abs_change": float(
                    fold_subset["median_standardized_abs_change"].median()
                ),
                "within_to_total_variance_ratio": float(
                    fold_subset["within_to_total_variance_ratio"].median()
                ),
                "shortcut_risk": shortcut_row["shortcut_risk"],
                "shortcut_static_geometry_spearman": shortcut_row["static_median_abs_spearman"],
                "shortcut_change_geometry_spearman": shortcut_row["change_median_abs_spearman"],
                "mri_observable": observable_row["mri_observable"],
                "observability_gate": observable_row["gate"],
                "g3_static_decodability_spearman": (
                    float(g3_static["macro_mean_spearman"].iloc[0]) if not g3_static.empty else math.nan
                ),
                "g3_delta_decodability_spearman": (
                    float(g3_delta["macro_mean_spearman"].iloc[0]) if not g3_delta.empty else math.nan
                ),
                "overall": overall[candidate],
                "decision_rationale": rationales[candidate],
            }
        )
    return pd.DataFrame(rows)


def configure_plots() -> None:
    cjk_font = Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Medium.ttc")
    if cjk_font.is_file():
        font_manager.fontManager.addfont(str(cjk_font))
    sns.set_theme(style="whitegrid", context="talk")
    plt.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": ["Noto Sans CJK JP", "DejaVu Sans"],
            "axes.unicode_minus": False,
            "figure.dpi": 120,
            "savefig.dpi": 220,
            "axes.titlesize": 14,
            "axes.labelsize": 11,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
        }
    )


def save_figure(fig: plt.Figure, name: str) -> None:
    fig.savefig(FIGURES / name, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def make_figures(
    coverage: pd.DataFrame,
    static: pd.DataFrame,
    delta: pd.DataFrame,
    pair_static: pd.DataFrame,
    pair_delta: pd.DataFrame,
    standardized_change: pd.DataFrame,
    variability: pd.DataFrame,
    residual: pd.DataFrame,
    fold_metrics: pd.DataFrame,
    decision: pd.DataFrame,
) -> None:
    configure_plots()

    # 1. Coverage/missingness
    coverage_columns = [
        "T0_available_overlap",
        "T1_available_overlap",
        "T2_available_overlap",
        "T3_available_overlap",
        "T0_to_T1_paired_overlap",
        "T1_to_T2_paired_overlap",
        "T2_to_T3_paired_overlap",
    ]
    coverage_labels = ["T0", "T1", "T2", "T3", "Δ01", "Δ12", "Δ23"]
    cov = coverage.set_index("feature").loc[list(ALL_FEATURES), coverage_columns] / 375.0 * 100.0
    cov.columns = coverage_labels
    fig, ax = plt.subplots(figsize=(9, 4.5))
    sns.heatmap(cov, annot=True, fmt=".1f", cmap="Blues", vmin=0, vmax=100, cbar_kws={"label": "%"}, ax=ax)
    ax.set_title("Strict MRI-overlap measurement coverage（n=375）")
    ax.set_xlabel("Visit / adjacent transition")
    ax.set_ylabel("Feature")
    save_figure(fig, "01_feature_coverage.png")

    # 2. Static FTV redundancy: cross-fold median in each visit.
    formal_static = static[static["selection_scope"]].copy()
    static_heat = formal_static.pivot_table(
        index="candidate", columns="visit", values="abs_spearman", aggfunc="median"
    ).reindex(index=CANDIDATES, columns=VISITS)
    fig, ax = plt.subplots(figsize=(7, 4.5))
    sns.heatmap(static_heat, annot=True, fmt=".3f", cmap="mako", vmin=0, vmax=1, ax=ax)
    ax.set_title("Static redundancy with FTV\nmedian |Spearman ρ| across fold-train analyses")
    ax.set_xlabel("Visit")
    ax.set_ylabel("Candidate")
    save_figure(fig, "02_static_ftv_redundancy.png")

    # 3. Delta FTV redundancy.
    formal_delta = delta[delta["selection_scope"]].copy()
    delta_heat = formal_delta.pivot_table(
        index="candidate", columns="transition", values="abs_spearman", aggfunc="median"
    ).reindex(index=CANDIDATES, columns=TRANSITIONS)
    fig, ax = plt.subplots(figsize=(7, 4.5))
    sns.heatmap(delta_heat, annot=True, fmt=".3f", cmap="mako", vmin=0, vmax=1, ax=ax)
    ax.set_title("Longitudinal redundancy with raw ΔFTV\nmedian |Spearman ρ| across fold-train analyses")
    ax.set_xlabel("Adjacent transition")
    ax.set_ylabel("Candidate")
    save_figure(fig, "03_delta_ftv_redundancy.png")

    # 4. Pairwise static correlations.
    fig, axes = plt.subplots(1, 4, figsize=(17, 4.2), constrained_layout=True)
    for ax, visit in zip(axes, VISITS):
        subset = pair_static[pair_static["visit"] == visit].set_index("feature_row").loc[list(ALL_FEATURES), list(ALL_FEATURES)]
        sns.heatmap(subset, annot=True, fmt=".2f", cmap="vlag", center=0, vmin=-1, vmax=1, cbar=visit == "T3", ax=ax)
        ax.set_title(visit)
        ax.set_xlabel("")
        ax.set_ylabel("")
    fig.suptitle("Pairwise static Spearman correlation（384-patient workbook）", y=1.04)
    save_figure(fig, "04_pairwise_static_correlation.png")

    # 5. Pairwise delta correlations.
    fig, axes = plt.subplots(1, 3, figsize=(13, 4.2), constrained_layout=True)
    for ax, transition in zip(axes, TRANSITIONS):
        subset = pair_delta[pair_delta["transition"] == transition].set_index("feature_row").loc[list(ALL_FEATURES), list(ALL_FEATURES)]
        sns.heatmap(subset, annot=True, fmt=".2f", cmap="vlag", center=0, vmin=-1, vmax=1, cbar=transition == "T2→T3", ax=ax)
        ax.set_title(transition)
        ax.set_xlabel("")
        ax.set_ylabel("")
    fig.suptitle("Pairwise raw-delta Spearman correlation（384-patient workbook）", y=1.04)
    save_figure(fig, "05_pairwise_delta_correlation.png")

    # 6. Longitudinal standardized change distributions.
    plot_change = standardized_change[standardized_change["feature"].isin(CANDIDATES)].copy()
    fig, ax = plt.subplots(figsize=(10, 5.5))
    sns.boxplot(
        data=plot_change,
        x="transition",
        y="standardized_abs_change",
        hue="feature",
        hue_order=CANDIDATES,
        showfliers=False,
        ax=ax,
    )
    ax.set_ylim(0, float(np.quantile(plot_change["standardized_abs_change"], 0.98)))
    ax.set_title("Longitudinal |ΔX| / start-visit IQR（outliers retained in metrics）")
    ax.set_xlabel("Transition")
    ax.set_ylabel("Standardized absolute change")
    ax.legend(title="Candidate", ncol=3)
    save_figure(fig, "06_longitudinal_standardized_change.png")

    # 7. Within- vs between-patient variance.
    var = (
        variability[(variability["fold"] == -1)]
        .drop_duplicates("feature")
        .set_index("feature")
        .loc[list(ALL_FEATURES)]
    )
    variance_plot = var[["within_to_total_variance_ratio", "between_patient_variance_fraction"]].rename(
        columns={
            "within_to_total_variance_ratio": "Within-patient",
            "between_patient_variance_fraction": "Between-patient",
        }
    )
    fig, ax = plt.subplots(figsize=(8, 5))
    variance_plot.plot(kind="bar", stacked=True, color=["#4c78a8", "#f2cf5b"], ax=ax)
    ax.set_ylim(0, 1)
    ax.set_ylabel("Fraction of total variance")
    ax.set_xlabel("Feature")
    ax.set_title("Within-patient vs between-patient variance")
    ax.legend(loc="upper right")
    ax.tick_params(axis="x", rotation=0)
    save_figure(fig, "07_within_between_variance.png")

    # 8. Residual information.
    resid_plot = (
        residual.groupby(["candidate", "task"], as_index=False)[
            ["residual_variance_ratio", "residual_iqr_ratio"]
        ]
        .median()
        .melt(
            id_vars=["candidate", "task"],
            value_vars=["residual_variance_ratio", "residual_iqr_ratio"],
            var_name="metric",
            value_name="ratio",
        )
    )
    resid_plot["panel"] = resid_plot["task"] + " / " + resid_plot["metric"].map(
        {"residual_variance_ratio": "variance", "residual_iqr_ratio": "IQR"}
    )
    fig, ax = plt.subplots(figsize=(10, 5.5))
    sns.barplot(data=resid_plot, x="panel", y="ratio", hue="candidate", hue_order=CANDIDATES, ax=ax)
    ax.axhline(1.0, color="black", linestyle="--", linewidth=1)
    ax.set_title("Residual information beyond FTV（median across folds/cells）")
    ax.set_xlabel("")
    ax.set_ylabel("Residual / original ratio")
    ax.tick_params(axis="x", rotation=15)
    ax.legend(title="Candidate", loc="upper left", bbox_to_anchor=(1.01, 1.0))
    save_figure(fig, "08_residual_information.png")

    # 9. Fold consistency.
    fig, axes = plt.subplots(2, 2, figsize=(12, 8.5))
    fold_panels = [
        ("static_median_abs_spearman_with_ftv", "Static |ρ(X,FTV)|", False),
        ("delta_median_abs_spearman_with_delta_ftv", "Delta |ρ(ΔX,ΔFTV)|", False),
        ("delta_median_residual_variance_ratio", "Delta residual variance ratio", True),
        ("within_to_total_variance_ratio", "Within / total variance", True),
    ]
    palette = dict(zip(CANDIDATES, sns.color_palette("Set2", n_colors=3)))
    for ax, (column, title, higher_better) in zip(axes.flat, fold_panels):
        sns.lineplot(
            data=fold_metrics,
            x="fold",
            y=column,
            hue="candidate",
            hue_order=CANDIDATES,
            marker="o",
            palette=palette,
            ax=ax,
        )
        ax.set_title(title + ("（higher=more）" if higher_better else "（lower=less redundant）"))
        ax.set_xticks(range(5))
        ax.set_ylabel("")
        ax.legend_.remove()
    handles, labels = axes.flat[0].get_legend_handles_labels()
    fig.tight_layout(rect=(0, 0.12, 1, 1), h_pad=2.0, w_pad=1.5)
    fig.legend(handles, labels, title="Candidate", loc="lower center", ncol=3, bbox_to_anchor=(0.5, 0.01))
    save_figure(fig, "09_candidate_fold_consistency.png")

    # 10. Decision matrix without an aggregate weighted score.
    rating_rows = []
    for row in decision.itertuples(index=False):
        rating_rows.append(
            {
                "candidate": row.candidate,
                "Coverage": 2,
                "Static complement": 2 if row.static_ftv_redundancy < 0.4 else 1 if row.static_ftv_redundancy < 0.7 else 0,
                "Delta complement": 2 if row.delta_ftv_redundancy < 0.4 else 1 if row.delta_ftv_redundancy < 0.7 else 0,
                "Residual info": 2 if min(row.static_residual_variance_ratio, row.delta_residual_variance_ratio) >= 0.8 else 1,
                "Longitudinal variation": 2 if row.within_to_total_variance_ratio >= 0.30 else 1,
                "Shortcut safety": (
                    1 if row.candidate == "BPE" else {"LOW": 2, "MODERATE": 1, "HIGH": 0}[row.shortcut_risk]
                ),
                "Input observable": {
                    "NO_WITH_CURRENT_INPUT": 0,
                    "YES_WITH_CAVEAT": 1,
                    "YES": 2,
                }[row.mri_observable],
            }
        )
    rating = pd.DataFrame(rating_rows).set_index("candidate")
    annotations = rating.replace({0: "Gate/risk", 1: "Caveat", 2: "Favorable"})
    fig, ax = plt.subplots(figsize=(12, 4.5))
    sns.heatmap(
        rating,
        annot=annotations,
        fmt="",
        cmap=sns.color_palette(["#d95f5f", "#f2cf5b", "#59a14f"], as_cmap=True),
        vmin=0,
        vmax=2,
        cbar=False,
        linewidths=1,
        linecolor="white",
        ax=ax,
    )
    ax.set_title("Multi-criteria decision matrix（no weighted aggregate score）")
    ax.set_xlabel("")
    ax.set_ylabel("Candidate")
    save_figure(fig, "10_final_decision_matrix.png")


def write_schema_report(
    schema: pd.DataFrame,
    metadata: dict[str, Any],
    coverage: pd.DataFrame,
    mapping_summary: dict[str, Any],
) -> None:
    compact = schema[
        [
            "column",
            "dtype",
            "role",
            "feature",
            "visit_or_interval",
            "n_missing",
            "n_unique",
            "n_zero",
            "min",
            "median",
            "max",
        ]
    ]
    report = f"""# I-SPY2 Multi-feature MRI NACT 表结构审计

## 结论

真实读取的工作簿只有一个 sheet：`{SHEET}`，数据维度为 384×29（Excel used range `{metadata['dimension']}`）。384 行对应 384 个唯一六位 ClinicalTrialSubjectID；缺失单元格、无限值、重复患者和重复整行均为 0。工作簿 SHA-256 为 `{sha256(WORKBOOK)}`。

表为一患者一行的 wide longitudinal structure，没有显式 `visit` 列。FTV 的原始列名特殊：`VOLUME_TUM_BLU_V10/V20/V30/V40` 分别对应 T0/T1/T2/T3；SPH、LD、BPE 直接在列名中编码 T0–T3。

## 工作簿结构核验

- Sheet names：`{metadata['sheet_names']}`；sheet state `{metadata['sheet_state']}`；无隐藏补充数据表。
- 行/列：表头 1 行 + 384 数据行，29 列。
- 合并单元格：{metadata['merged_ranges']}；公式单元格：{metadata['formula_cells']}。
- 重复患者：{metadata['duplicate_patient_ids']}；重复整行：{metadata['duplicate_rows']}；忽略 ID 后重复 measurement vector：{metadata['duplicate_rows_without_id']}。
- 29 列全部 0% missing；ID 为 `int64`，其余 28 列为 `float64`。
- 工作簿不含 pCR、molecular subtype、treatment、age 或其他 clinical outcome/metadata 字段。

## 真正的 longitudinal measurement families

| Family | 原始四访列 | 角色 | 定义/单位 |
|---|---|---|---|
| FTV | `VOLUME_TUM_BLU_V10`–`V40` | reference target | 功能性肿瘤体积；cc |
| LD | `LD_T0`–`LD_T3` | formal candidate | 影像报告最长径；工作簿未明示单位 |
| SPH | `SPHERICITY_T0`–`T3` | formal candidate | 等体积球表面积 / 3D FTV mask 表面积；无量纲 |
| BPE | `BPE_5slice_mean_T0`–`T3` | formal candidate | 对侧乳腺中央 5 层纤维腺体 mean PE |

因此 formal candidate pool 恰为 LD、SPH、BPE；FTV 是 reference。论文方法用于补足工作簿未写出的 measurement 定义：{PRIMARY_PAPER}。

## 派生列验证

12 个 `*_pch_T0_Tk` 列均为已经物化的 baseline-relative 百分比，逐行验证公式 `100 × (X_Tk−X_T0)/X_T0`；各 family 最大绝对误差为 `{json.dumps(metadata['pch_max_abs_error'], ensure_ascii=False)}`。它们不是新的独立 measurement，也不是 T1→T2/T2→T3 相邻变化，因此不进入 candidate pool；正式相邻 Δ 从四访 raw columns 重新计算。

## Coverage 与零值质量信号

{markdown_table(coverage[['feature','workbook_total_patients','strict_mri_overlap_patients','T0_available_workbook','T1_available_workbook','T2_available_workbook','T3_available_workbook','complete_4visit_workbook','missing_pct_workbook']])}

所有四类 measurement 在该 complete-case workbook 中均为 384/384 四访完整；这不能外推为原始 808 cohort 无缺失。零值不被改写为 NA：LD T0/T1/T2/T3 的零值数为 0/6/65/128，BPE 为 0/0/1/4。LD 后期存在明显 floor effect，可能包括病灶消失/不可测或编码下限；源表不能区分其语义，本 screening 保留并标记，不任意删除。

## Patient mapping

仅用 canonical patient ID 严格 regex 后缀、clinical ID 与 workbook 六位 ID 三者等值匹配，不做 fuzzy matching：workbook {mapping_summary['workbook_patient_count']} 人、MRI cohort {mapping_summary['mri_cohort_patient_count']} 人、交集 {mapping_summary['strict_overlap_count']} 人、workbook-only {mapping_summary['workbook_only_count']} 人、MRI-only {mapping_summary['mri_only_count']} 人。workbook-only IDs：`{', '.join(mapping_summary['workbook_only_ids'])}`。

## 逐列 schema

{markdown_table(compact)}
"""
    (REPORTS / "table_schema_report.md").write_text(report, encoding="utf-8")


def report_summary_tables(
    static: pd.DataFrame,
    delta: pd.DataFrame,
    residual: pd.DataFrame,
    variability: pd.DataFrame,
    distribution: pd.DataFrame,
    fold_metrics: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    static_summary = (
        static[static["selection_scope"]]
        .pivot_table(index="candidate", columns="visit", values="abs_spearman", aggfunc="median")
        .reindex(index=CANDIDATES, columns=VISITS)
        .reset_index()
    )
    static_formal = (
        fold_metrics.groupby("candidate")["static_median_abs_spearman_with_ftv"]
        .median()
        .to_dict()
    )
    static_summary["跨访视/折中位数"] = static_summary["candidate"].map(static_formal)
    static_summary["分类"] = static_summary["跨访视/折中位数"].map(redundancy_class)

    delta_summary = (
        delta[delta["selection_scope"]]
        .pivot_table(index="candidate", columns="transition", values="abs_spearman", aggfunc="median")
        .reindex(index=CANDIDATES, columns=TRANSITIONS)
        .reset_index()
    )
    delta_formal = (
        fold_metrics.groupby("candidate")["delta_median_abs_spearman_with_delta_ftv"]
        .median()
        .to_dict()
    )
    delta_summary["跨transition/折中位数"] = delta_summary["candidate"].map(delta_formal)
    delta_summary["分类"] = delta_summary["跨transition/折中位数"].map(redundancy_class)

    residual_summary = (
        residual.groupby(["candidate", "task"], as_index=False)[
            ["residual_variance_ratio", "residual_iqr_ratio"]
        ]
        .median()
        .pivot(index="candidate", columns="task")
    )
    residual_summary.columns = [f"{task}_{metric}" for metric, task in residual_summary.columns]
    residual_summary = residual_summary.reindex(CANDIDATES).reset_index()

    variation_summary = (
        variability[(variability["selection_scope"]) & (variability["feature"].isin(CANDIDATES))]
        .groupby("feature", as_index=False)
        .agg(
            median_standardized_abs_change=("median_standardized_abs_change", "median"),
            median_abs_delta=("median_abs_delta", "median"),
            within_to_total_variance_ratio=("within_to_total_variance_ratio", "median"),
            near_zero_fraction=("proportion_near_zero", "median"),
        )
        .rename(columns={"feature": "candidate"})
        .set_index("candidate")
        .reindex(CANDIDATES)
        .reset_index()
    )

    quality_summary = distribution[
        distribution["feature"].isin(CANDIDATES)
    ][
        [
            "feature",
            "visit",
            "median",
            "iqr",
            "n_zero",
            "zero_fraction",
            "heavy_tail_flag",
            "floor_effect_flag",
            "near_zero_variance_flag",
        ]
    ]

    stability = (
        fold_metrics.groupby("candidate", as_index=False)
        .agg(
            static_rho_min=("static_median_abs_spearman_with_ftv", "min"),
            static_rho_max=("static_median_abs_spearman_with_ftv", "max"),
            delta_rho_min=("delta_median_abs_spearman_with_delta_ftv", "min"),
            delta_rho_max=("delta_median_abs_spearman_with_delta_ftv", "max"),
            residual_delta_min=("delta_median_residual_variance_ratio", "min"),
            residual_delta_max=("delta_median_residual_variance_ratio", "max"),
            within_ratio_min=("within_to_total_variance_ratio", "min"),
            within_ratio_max=("within_to_total_variance_ratio", "max"),
        )
        .set_index("candidate")
        .reindex(CANDIDATES)
        .reset_index()
    )
    return {
        "static": static_summary,
        "delta": delta_summary,
        "residual": residual_summary,
        "variation": variation_summary,
        "quality": quality_summary,
        "stability": stability,
    }


def write_final_report(
    schema_metadata: dict[str, Any],
    mapping_summary: dict[str, Any],
    coverage: pd.DataFrame,
    static: pd.DataFrame,
    delta: pd.DataFrame,
    residual: pd.DataFrame,
    variability: pd.DataFrame,
    distribution: pd.DataFrame,
    pair_static: pd.DataFrame,
    pair_delta: pd.DataFrame,
    shortcut: pd.DataFrame,
    observability: pd.DataFrame,
    decodability: pd.DataFrame,
    fold_metrics: pd.DataFrame,
    decision: pd.DataFrame,
) -> None:
    tables = report_summary_tables(static, delta, residual, variability, distribution, fold_metrics)
    decision_display = decision[
        [
            "candidate",
            "coverage",
            "static_ftv_redundancy",
            "delta_ftv_redundancy",
            "static_residual_variance_ratio",
            "delta_residual_variance_ratio",
            "within_to_total_variance_ratio",
            "shortcut_risk",
            "mri_observable",
            "overall",
        ]
    ].rename(
        columns={
            "candidate": "Candidate",
            "coverage": "Coverage",
            "static_ftv_redundancy": "Static FTV redundancy",
            "delta_ftv_redundancy": "ΔFTV redundancy",
            "static_residual_variance_ratio": "Static residual info",
            "delta_residual_variance_ratio": "Δ residual info",
            "within_to_total_variance_ratio": "Longitudinal variation",
            "shortcut_risk": "Shortcut risk",
            "mri_observable": "MRI observable",
            "overall": "Overall",
        }
    )

    static_winner = tables["static"].sort_values("跨访视/折中位数").iloc[0]
    delta_winner = tables["delta"].sort_values("跨transition/折中位数").iloc[0]
    residual_rank = tables["residual"].copy()
    residual_rank["combined"] = residual_rank[
        ["static_residual_variance_ratio", "delta_residual_variance_ratio"]
    ].median(axis=1)
    residual_winner = residual_rank.sort_values("combined", ascending=False).iloc[0]
    variation_winner = tables["variation"].sort_values(
        ["within_to_total_variance_ratio", "median_standardized_abs_change"], ascending=False
    ).iloc[0]
    amplitude_winner = tables["variation"].sort_values(
        "median_standardized_abs_change", ascending=False
    ).iloc[0]
    ld_variation = tables["variation"].set_index("candidate").loc["LD"]
    sph_variation = tables["variation"].set_index("candidate").loc["SPH"]

    shortcut_display = shortcut[
        [
            "feature",
            "shortcut_risk",
            "static_median_abs_spearman",
            "static_median_r2",
            "change_median_abs_spearman",
            "change_median_r2",
            "formal_ci_status",
            "interpretation",
        ]
    ]
    observability_display = observability[
        ["feature", "required_imaging_region", "mri_observable", "gate", "rationale"]
    ]
    decodability_display = decodability[
        [
            "feature",
            "model",
            "task",
            "macro_mean_spearman",
            "macro_mean_r2",
            "cell_values",
        ]
    ]

    report = f"""# Radiomics Target Screening 最终报告

## 1. 最终结论

在**不改变当前 G3 input contract**（DCE7 lesion/ROI-centered crop → encoder → GAP，严格无 mask channel），并明确优先“shortcut safety + 当前可执行性”的 operational preference 下，唯一第一 pilot 推荐是 **LD（longest diameter）**；第二候选是 **SPH（sphericity）**。这是一个透明的 pragmatic/conditional adjudication，不是统计数据唯一推出的全序：LD 与 SPH 在定量维度上均为 Pareto non-dominated。BPE 的 **static** FTV 互补性最强，但其测量需要对侧乳腺中央 5 层纤维腺体，当前 lesion-centered crop 不保证包含所需解剖区域，因此被 architecture observability gate 淘汰；动态 Δ 互补性最强的是 SPH。

LD 并非完美候选：它仍是 tumor burden measure，static 与 FTV 中等冗余，且 T2/T3 出现明显零值 floor。它胜出不是因为单一综合分数，而是因为在上述 operational preference 下，LD 同时具有完整 coverage、当前输入的部分可观察证据、稳定纵向变化、较高 FTV residual information、仅中等而非高 mask-geometry 风险。既有 strict-DCE7 probe 的 ΔLD Spearman 点估计高于 ΔSPH，但二者 target transform 不同且没有 paired cross-target CI，因此只作弱 feasibility 旁证，不构成正式优效检验。固定 crop 是否逐例容纳完整最长径尚未被证明，因此 crop-containment audit 是进入 pilot 前的必要 gate。

## 2. Outcome-free screening 边界

本 screening 的目标不是寻找与 pCR correlation 最大的 biomarker，而是寻找一个与 FTV complementary 的 imaging-response axis，供下一阶段 factorized/multi-dimensional response state 使用。正式 selection 代码只读取 fold manifest 的 `patient_id/fold/split` 三列，**没有读取 `label_pcr`**；没有使用 pCR AUROC、pCR association、treatment response label、molecular subtype 或任何 clinical outcome。下一阶段才能在严格 held-out 条件下评估额外 response axis 是否改善 pCR prediction，从而避免 target-selection bias。

## 3. Excel 真实结构

- 文件：`{WORKBOOK}`；SHA-256 `{sha256(WORKBOOK)}`。
- Sheet：仅 `{SHEET}`；data shape 384×29，Excel used range `{schema_metadata['dimension']}`。
- 384 个唯一六位患者 ID；所有 29 列 0% missing；0 个重复患者、0 个重复整行、0 个 non-finite。
- 表为 wide longitudinal structure；FTV `V10/V20/V30/V40` 映射 T0/T1/T2/T3。
- 12 个 `pch` 字段全部逐行验证为相对 T0 百分比派生量，不是新的 measurement，也不是相邻 Δ。

详细逐列检查见 [table_schema_report.md](table_schema_report.md) 和 `../metrics/table_schema.csv`。

## 4. Measurement 定义与候选池

| Measurement | 角色 | 定义 | 当前 target 含义 |
|---|---|---|---|
| FTV | reference | 满足 PE/SER 阈值的增强组织功能性肿瘤体积 | 主 tumor burden response axis |
| LD | formal candidate | site radiologist MRI report 中的肿瘤最长径 | 线性病灶范围/burden axis |
| SPH | formal candidate | 等体积球表面积除以 3D FTV tumor mask 表面积 | tumor morphology/compactness axis |
| BPE | formal candidate | 对侧乳腺中央连续 5 层纤维腺体的平均早期 PE | normal-tissue physiology axis |

定义来源为同一 384 人表对应的原始论文方法：{PRIMARY_PAPER}。工作簿中没有其他真正的 longitudinal MRI-derived quantitative measurement；`pch` 列只作一致性核验。

## 5. Patient coverage 与严格 mapping

{markdown_table(coverage[['feature','workbook_total_patients','strict_mri_overlap_patients','T0_available_overlap','T1_available_overlap','T2_available_overlap','T3_available_overlap','complete_4visit_overlap','missing_pct_overlap']])}

工作簿 384 人、MRI cohort 808 人，严格 overlap 375 人，workbook-only 9 人、MRI-only 433 人。只允许 `^(ISPY2-|ACRIN-6698-)######` 后缀、clinical ID 与 workbook 六位 ID 三者精确等值，不做 fuzzy matching。workbook-only IDs 为 `{', '.join(mapping_summary['workbook_only_ids'])}`。

## 6. Fold-safe protocol

锁定 seed-2026 manifest SHA-256 `{sha256(FOLD_MANIFEST)}`。每个 outer fold 只用显式 `split=train` 的 measurement-overlap 患者计算正式指标；五折 n 分别为 {mapping_summary['fold_train_overlap_counts']}. validation/test 没有进入 ranking。全 384 人结果仅作 descriptive summary。

相关性主指标为 absolute Spearman，Pearson 辅助；raw longitudinal change 定义为相邻 endpoint 的 `X_end-X_start`。Residual analysis 在各 fold/cell 内使用简单 `StandardScaler + Ridge(alpha=1)`；static predictor 为 `log1p(FTV)`，LD target 使用 `log1p`，SPH/BPE 为 identity；raw delta 模型不作 log 变换。该同一训练集拟合/描述关系用于量化可被 FTV 解释的线性信息，不用于泛化性能宣称。

## 7. Static FTV redundancy

{markdown_table(tables['static'])}

Static 最不冗余的是 **{static_winner['candidate']}**（跨 fold/visit 中位 |ρ|={fmt(static_winner['跨访视/折中位数'])}）。BPE 始终只有弱 static FTV correlation。LD 属中等冗余，符合其线性 tumor burden 含义。SPH 与 FTV 呈强负相关（主表 signed ρ 为负），因此 absolute redundancy 反而高于 LD；到 T3 尤其明显。相关方向不影响“是否复制 FTV 信息”的判定，所以正式指标取绝对值。

![Static FTV redundancy](../figures/02_static_ftv_redundancy.png)

## 8. Longitudinal raw-Δ redundancy

{markdown_table(tables['delta'])}

Raw change 与 ΔFTV 最不冗余的是 **{delta_winner['candidate']}**（跨 fold/transition 中位 |ρ|={fmt(delta_winner['跨transition/折中位数'])}）。SPH 在 T0→T1/T1→T2 基本独立于 ΔFTV，但 T2→T3 冗余上升；LD 的 ΔFTV redundancy 为低至中等；BPE 为低冗余，但不如其 static 结果极端。

![Delta FTV redundancy](../figures/03_delta_ftv_redundancy.png)

## 9. Residual information beyond FTV

{markdown_table(tables['residual'])}

按实现中预先固定的 simple-transform static/delta residual variance ratio 中位描述，保留最多信息的是 **{residual_winner['candidate']}**（中位 ratio={fmt(residual_winner['combined'])}）。本节表格先把所有 fold×cell 直接取中位；Decision Matrix 则先在每个 fold 内跨 cell 取中位、再跨 fold 取中位，因此小数可略有差异，但候选排序不变。三个候选都不是由简单单变量 FTV relation 完全决定；BPE static 最接近完全残留，SPH delta 也几乎完全残留。LD 虽为另一种 burden measure，仍保留大部分方差/IQR。这些 ratio 是各 feature 自身 transform space 内的保留比例（LD static 使用 log1p，SPH/BPE 使用 identity），适合判断“是否容易由 FTV 解释”，但不能把跨 feature 的小数差异当成严格同量纲效应量。

![Residual information](../figures/08_residual_information.png)

## 10. Longitudinal responsiveness

{markdown_table(tables['variation'])}

纵向响应的答案取决于 metric。按本任务的主摘要 within-patient / total variance，**{variation_winner['candidate']}** 最高（{fmt(variation_winner['within_to_total_variance_ratio'])}）；按 standardized |Δ|，**{amplitude_winner['candidate']}** 最高（{fmt(amplitude_winner['median_standardized_abs_change'])}；LD={fmt(ld_variation['median_standardized_abs_change'])}，SPH={fmt(sph_variation['median_standardized_abs_change'])}）。因此不存在不加限定的“最强变化”候选；必答 F 按 within/total 口径回答 LD。三个候选都不是 near-constant patient trait：within/total ratio 均为实质正值，相邻变化的 near-zero 比例低。LD 后期 floor 会把部分变化压到 0；SPH/BPE 也有明确治疗期变化。

![Standardized change](../figures/06_longitudinal_standardized_change.png)

![Within vs between](../figures/07_within_between_variance.png)

## 11. Dynamic range 与 measurement quality

{markdown_table(tables['quality'])}

所有 feature×visit 均非 near-constant，未删除任何 outlier。FTV heavy-tail 最明显；LD 在 T2/T3 分别有 65/128 个零值，T3 达到显著 floor effect。BPE T2/T3 有 1/4 个零值。源表不能区分“真实完全消失、不可测、分割失败或编码下限”，因此这些值保留原样，并在推荐中把 LD floor 作为风险而非悄然清洗。

## 12. Candidate–candidate redundancy

完整 static 和 raw-delta Spearman matrices 分别保存在 `pairwise_static_spearman.csv` 与 `pairwise_delta_spearman.csv`，按每个 visit/transition 分块。主要结构是：LD 与 FTV 形成 tumor burden 轴；SPH 与 FTV/LD 多为负相关的 compactness/morphology 轴；BPE 与病灶 features 的 static correlation 最弱，构成 normal-tissue physiology 轴。

![Pairwise static](../figures/04_pairwise_static_correlation.png)

![Pairwise delta](../figures/05_pairwise_delta_correlation.png)

## 13. Shortcut / mask-geometry audit

{markdown_table(shortcut_display)}

9-D geometry control 包括 ROI voxel volume、normalized bbox extents/diagonal 与 centroid。SPH 为 **HIGH**：它直接由 3D FTV mask 表面积定义，static geometry predictability 高，后两段 change 也明显。LD 为 **MODERATE**：static 可由 lesion geometry 中等预测，但历史 log-difference coordinate 的 change R² 弱。BPE 的观察点估计为 **LOW**，但没有正式 BPE bootstrap CI，不能写成已完全排除 shortcut。

重要边界：这些 control 说明 target 对显式 mask geometry 的依赖，**不代表 G3 已经读取 mask**。G3 代码硬性排除了 mask channel；仍存在 ROI-centered crop 的定位先验，但没有直接 voxel-count/geometry scalar route。历史 audit 的 ΔFTV/ΔLD 使用 log difference，而本 screening 使用任务指定的 raw difference，两者只作旁证，不直接数值等同。

## 14. Current MRI input observability gate

{markdown_table(observability_display)}

BPE 的 static 统计互补性不能覆盖 architecture mismatch：对侧乳腺中央 tissue 不在 lesion-centered crop 的保证区域内。因此 BPE 是 **STATICALLY ATTRACTIVE BUT INPUT-MISMATCHED**，在当前 input contract 下不能推荐。LD/SPH 的 ROI-centered crop 含病灶定位区域，且 strict-DCE7 static probe 提供弱正 proxy signal，因此仅作 **conditional pass**；现有契约没有逐例证明完整最长径或完整 3D 表面未被固定 crop 截断。SPH 还因没有显式 mask、GAP 弱化空间结构而带有更强 caveat。

## 15. Prior frozen representation decodability

{markdown_table(decodability_display)}

Strict-DCE7 G3 对 LD/SPH 的 static rank signal 均为弱至中等；ΔLD Spearman 点估计为弱正，ΔSPH 接近零且 R² 为负。由于两者历史 target transform 不同且没有 paired cross-target CI，这一差异不是正式的跨 target 优效检验。BPE 没有同 contract 的 DGRS probe，记为 UNKNOWN；含 mask 的旧 OSRA 弱 signal 不能越过 observability gate。Decodability 仅作 feasibility 旁证，不是“选当前最容易预测的 target”的单一规则。

## 16. Cross-fold consistency

{markdown_table(tables['stability'])}

三个候选的相对结论在五个 fold-train 子集中没有翻转：BPE 持续 static 最不冗余；SPH 持续 raw-delta 最不冗余；LD 持续表现出较强 within-patient variation。这里的五个 train sets 高度重叠，是 split sensitivity/consistency check，不是五个独立重复实验；正式 recommendation 取这种一致的 multi-criteria pattern，而非 held-out fold 表现。

![Fold consistency](../figures/09_candidate_fold_consistency.png)

## 17. Candidate Decision Matrix

{markdown_table(decision_display)}

![Decision matrix](../figures/10_final_decision_matrix.png)

决策采用 Pareto/multi-criteria gate，不存在人为权重相加得到的唯一分数。LD 与 SPH 在统计维度上均 Pareto non-dominated；为满足任务要求给出唯一第一候选，本报告显式用“优先降低 geometry shortcut、优先当前 strict-DCE7 可执行性”作 qualitative tiebreaker：

1. **LD — RECOMMENDED（conditional/pragmatic first）。** 在上述 tiebreaker 下，数据、纵向响应、部分可观察证据和 shortcut safety 的可行交集最好。它比 SPH 少一个 HIGH geometry-risk gate，比 BPE 少一个 input-mismatch hard failure。代价是与 FTV 中等 static redundancy、后期 floor，以及完整最长径可能超出固定 crop；下一阶段前必须完成 crop-containment audit。
2. **SPH — POSSIBLE SECOND CHOICE。** 它是最清晰的 morphology candidate，ΔSPH 与 ΔFTV 最互补且 residual 最大；但 SPH 的精确定义就是 FTV mask surface geometry，历史 strict-DCE7 Δ解码几乎为零，因此不作为第一候选。
3. **BPE — NOT RECOMMENDED WITH CURRENT INPUT。** 它在 static 统计上最 complementary，但所需对侧乳腺区域不在当前输入 contract；只有扩大/增加全乳或对侧乳腺输入后才值得重新评估。

## 18. LD / SPH / BPE 专项回答

1. **LD 是否主要只是另一种 tumor burden measure？** 是，static 与 FTV 中等相关，且二者均随病灶缩小；但 raw Δ 相关低于 static、Ridge residual 仍高，因此不是 FTV 的完全复制。
2. **SPH 是否增加 morphology dimension？** 定义与表格统计支持一个 morphology axis：它刻画等体积条件下的表面复杂度/球形度，raw Δ 与 ΔFTV 最不冗余；但当前 G3 的 longitudinal morphology decodability 尚未建立（ΔSPH ρ=0.034，R²=-0.028）。
3. **SPH 是否过度依赖 mask geometry？** 是。定义直接使用 3D FTV mask 表面积，9-D mask geometry 的 static predictability 高，shortcut 风险为 HIGH。
4. **BPE 是否统计上与 FTV 最互补？** Static 是；delta 也低冗余且 residual 很高，但 delta 最不冗余的是 SPH。
5. **当前 lesion-centered crop 能否真正观察 BPE？** 不能合理保证。BPE 需要对侧乳腺中央 5 层纤维腺体，当前 crop 是病灶中心固定小视野。
6. **当前 G3 input contract 下谁最适合作为第二 target？** LD。

## 19. 必答 A–J

- **A. 实际有哪些 longitudinal measurements？** FTV、LD、SPH、BPE 四类；每类 T0–T3。12 个 pch 列只是相对 T0 派生量。
- **B. FTV 外哪些可作候选？** LD、SPH、BPE；但 BPE 在当前 input contract 下不可执行。
- **C. Static 与 FTV 最不冗余？** {static_winner['candidate']}，formal median |ρ|={fmt(static_winner['跨访视/折中位数'])}。
- **D. Longitudinal change 与 ΔFTV 最不冗余？** {delta_winner['candidate']}，formal median |ρ|={fmt(delta_winner['跨transition/折中位数'])}。
- **E. 保留最多 FTV 外 residual information？** {residual_winner['candidate']}（按 static/delta residual variance ratio 的联合描述）；BPE 是 static residual 最强者，SPH 是 delta residual 最强者。
- **F. 最强且稳定的 within-patient variation？** 按 requested within/total 主摘要为 {variation_winner['candidate']}（{fmt(variation_winner['within_to_total_variance_ratio'])}）；按 standardized |Δ| 则为 {amplitude_winner['candidate']}（{fmt(amplitude_winner['median_standardized_abs_change'])}），因此结论需随 metric 限定。
- **G. 明显 mask/geometry shortcut risk？** SPH=HIGH；LD=MODERATE；BPE=LOW point estimate 但 CI 不足；FTV reference=HIGH。
- **H. 当前 DCE7 lesion-centered 下无法合理观察？** BPE。
- **I. 不改 input contract 的最推荐第二 target？** 在优先 shortcut safety 与当前可执行性的明确决策偏好下，LD（conditional first）。
- **J. 为什么？** 它通过 coverage/variation gate，具有当前输入的部分可观察证据，保留较多非 FTV 信息，mask shortcut 风险低于 SPH，且没有 BPE 的解剖 input mismatch；但需先用 crop-containment audit 确认完整最长径没有被系统性截断。

## 20. 下一阶段 dual-grounding pilot（本轮不运行）

建议固定同一 seed/fold/input/model contract，仅比较：

```text
H0: DCE7 → JEPA
H1: DCE7 → JEPA + FTV
H2: DCE7 → JEPA + FTV + LD
```

在运行 H2 前，先做不涉及结局的 crop-containment audit：定量检查每个 visit 的 reported LD/FTV support 是否完整落入固定 crop，并预注册 truncation failure gate；若 LD gate 失败，只有在 SPH 自身的完整 3D surface-containment gate 通过时才回到 SPH，否则应判定当前 input contract 无合适第二 target，而不是直接训练。通过后，主要 held-out 比较应覆盖 FTV static、raw/预定义 ΔFTV、LD static、ΔLD、image-only pCR，以及 optimization safety。LD target transform、zero/floor handling 与 loss scale 必须仅在 fold train 拟合；不得因 pCR 表现反向修改 LD target 选择。鉴于既有 G3 optimization 证据是 promising but unstable，H2 还应预注册 base degradation、gradient conflict、representation variance 和 failure-fold gate；但本 screening 没有训练 H0/H1/H2、没有改 lambda/JEPA/transition，也没有运行 optimization fix。

## 21. 局限性

- 工作簿是 384 人 complete-case subset；coverage 不能外推到所有 808 人。
- LD 单位与零值语义未在工作簿中明示；floor 需在 pilot 前由数据字典/生成者确认。
- LD/SPH 的 current-input observability 是基于 ROI-centered contract 与 frozen static probe 的 conditional inference；尚无逐例 crop-containment 证明。
- 五个 fold-train 集高度重叠，跨折范围只表示 split sensitivity，不是独立重复的不确定性区间。
- Static residual 的 candidate-specific transform 不同，且 Ridge 在同一 fold-train 样本拟合/描述；ratio 不应解释为跨 feature 严格可比的 out-of-sample performance。
- SPH/FTV/LD/BPE 的 shortcut/decodability 证据来自已完成实验，其 target transforms 与本轮 raw-delta 口径不完全相同。
- BPE 的 input mismatch 是基于论文定义与当前 crop contract 的 architecture gate；不是宣称影像中不存在任何 proxy signal。
- Ridge residual 是简单线性/轻变换关系审计，不是条件互信息估计，也不支持因果解释。
- `verification.json` 仅核验产物结构、哈希、登记口径与内部一致性；`PASS` 不等于外部科学验证或临床效用证明。
"""
    (REPORTS / "final_report.md").write_text(report, encoding="utf-8")


def write_selection_json(
    mapping_summary: dict[str, Any], decision: pd.DataFrame, fold_metrics: pd.DataFrame
) -> None:
    summaries: dict[str, Any] = {}
    for candidate in CANDIDATES:
        row = decision[decision["candidate"] == candidate].iloc[0]
        folds = fold_metrics[fold_metrics["candidate"] == candidate]
        summaries[candidate] = {
            "overall": row["overall"],
            "coverage_fraction": row["coverage_fraction"],
            "static_ftv_redundancy_median_abs_spearman": row["static_ftv_redundancy"],
            "delta_ftv_redundancy_median_abs_spearman": row["delta_ftv_redundancy"],
            "static_residual_variance_ratio": row["static_residual_variance_ratio"],
            "delta_residual_variance_ratio": row["delta_residual_variance_ratio"],
            "within_to_total_variance_ratio": row["within_to_total_variance_ratio"],
            "shortcut_risk": row["shortcut_risk"],
            "mri_observable": row["mri_observable"],
            "observability_gate": row["observability_gate"],
            "fold_metric_ranges": {
                column: [float(folds[column].min()), float(folds[column].max())]
                for column in (
                    "static_median_abs_spearman_with_ftv",
                    "delta_median_abs_spearman_with_delta_ftv",
                    "delta_median_residual_variance_ratio",
                    "within_to_total_variance_ratio",
                )
            },
            "rationale": row["decision_rationale"],
        }
    payload = {
        "schema_version": 1,
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "selection_objective": "outcome-free complementary MRI response axis beyond FTV",
        "recommended_target": "LD",
        "recommended_target_full_name": "longest diameter",
        "recommendation_status": "conditional_pragmatic_first",
        "selection_is_statistically_unique": False,
        "pareto_nondominated_current_input_candidates": ["LD", "SPH"],
        "qualitative_tiebreaker": "prioritize lower geometry-shortcut risk and current strict-DCE7 feasibility",
        "second_choice": "SPH",
        "statistically_attractive_but_input_mismatched": "BPE",
        "no_model_training_performed": True,
        "pcr_read_or_used_for_selection": False,
        "formal_selection_scope": "each outer fold's explicit training patients only",
        "multi_criteria_not_single_weighted_score": True,
        "workbook": {
            "path": str(WORKBOOK),
            "sha256": sha256(WORKBOOK),
            "sheet": SHEET,
        },
        "fold_manifest": {
            "path": str(FOLD_MANIFEST),
            "sha256": sha256(FOLD_MANIFEST),
            "provenance_boundary": "locked internally validated candidate copy; native generation provenance unavailable",
        },
        "strict_overlap_mapping": {
            "path": str(OVERLAP),
            "sha256": sha256(OVERLAP),
            "validation": "patient set equals fold manifest; has_radiomics equals exact workbook/cohort ID intersection; match_rule fail-closed",
        },
        "cohort_mapping": mapping_summary,
        "candidate_summaries": summaries,
        "selection_rule": [
            "sufficient longitudinal coverage",
            "not near-constant",
            "not completely redundant with FTV or raw ΔFTV",
            "meaningful within-patient longitudinal variation",
            "reasonably observable from current DCE7 lesion-centered input",
            "shortcut risk does not dominate feasibility",
        ],
        "required_pre_pilot_gate": (
            "outcome-free per-visit crop-containment audit must show reported LD/lesion support is not systematically truncated; "
            "if LD fails, reconsider SPH only after its own full-3D-surface containment gate passes; "
            "otherwise conclude no suitable target under the current input contract"
        ),
        "stopping_condition": "target screening and recommendation complete; no H0/H1/H2 pilot run",
    }
    (METRICS / "final_target_selection.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def write_manifest(required_files: list[Path]) -> None:
    rows = []
    for path in required_files:
        rows.append(
            {
                "path": str(path.relative_to(EXP_ROOT)),
                "exists": path.is_file(),
                "size_bytes": path.stat().st_size if path.is_file() else 0,
                "sha256": sha256(path) if path.is_file() else "",
            }
        )
    manifest = pd.DataFrame(rows)
    save_csv(manifest, "output_manifest.csv")
    if not manifest["exists"].all() or not (manifest["size_bytes"] > 0).all():
        raise RuntimeError("required output 缺失或为空")


def main() -> None:
    for directory in (METRICS, FIGURES, REPORTS):
        directory.mkdir(parents=True, exist_ok=True)

    raw, _, schema_metadata = read_and_validate_workbook()
    schema = build_schema(raw, schema_metadata)
    measurements = measurement_table(raw)
    formal, folds, mapping_summary = strict_cohort_mapping(measurements)
    coverage = build_coverage(measurements, formal, mapping_summary)
    static, delta = redundancy_tables(measurements, formal, folds)
    residual = residual_information(formal, folds)
    variability, standardized_change = longitudinal_variability(measurements, formal, folds)
    distribution = distribution_summary(measurements)
    pair_static, pair_delta = pairwise_matrices(measurements)
    shortcut = shortcut_summary()
    observability = observability_summary()
    decodability = prior_decodability_summary()
    fold_metrics = fold_level_metrics(formal, folds, static, delta, residual, variability)
    decision = decision_matrix(coverage, fold_metrics, shortcut, observability, decodability)

    save_csv(schema, "table_schema.csv")
    save_csv(coverage, "candidate_coverage.csv")
    save_csv(static, "static_ftv_redundancy.csv")
    save_csv(delta, "delta_ftv_redundancy.csv")
    save_csv(residual, "residual_information.csv")
    save_csv(variability, "longitudinal_variability.csv")
    save_csv(distribution, "feature_distribution_summary.csv")
    save_csv(pair_static, "pairwise_static_spearman.csv")
    save_csv(pair_delta, "pairwise_delta_spearman.csv")
    save_csv(shortcut, "shortcut_risk_summary.csv")
    save_csv(observability, "observability_summary.csv")
    save_csv(decodability, "prior_decodability_summary.csv")
    save_csv(fold_metrics, "fold_level_candidate_metrics.csv")
    save_csv(decision, "candidate_decision_matrix.csv")

    write_schema_report(schema, schema_metadata, coverage, mapping_summary)
    make_figures(
        coverage,
        static,
        delta,
        pair_static,
        pair_delta,
        standardized_change,
        variability,
        residual,
        fold_metrics,
        decision,
    )
    write_final_report(
        schema_metadata,
        mapping_summary,
        coverage,
        static,
        delta,
        residual,
        variability,
        distribution,
        pair_static,
        pair_delta,
        shortcut,
        observability,
        decodability,
        fold_metrics,
        decision,
    )
    write_selection_json(mapping_summary, decision, fold_metrics)

    required = [
        METRICS / name
        for name in (
            "table_schema.csv",
            "candidate_coverage.csv",
            "static_ftv_redundancy.csv",
            "delta_ftv_redundancy.csv",
            "residual_information.csv",
            "longitudinal_variability.csv",
            "feature_distribution_summary.csv",
            "pairwise_static_spearman.csv",
            "pairwise_delta_spearman.csv",
            "shortcut_risk_summary.csv",
            "observability_summary.csv",
            "prior_decodability_summary.csv",
            "fold_level_candidate_metrics.csv",
            "candidate_decision_matrix.csv",
            "final_target_selection.json",
        )
    ]
    required.extend(
        FIGURES / name
        for name in (
            "01_feature_coverage.png",
            "02_static_ftv_redundancy.png",
            "03_delta_ftv_redundancy.png",
            "04_pairwise_static_correlation.png",
            "05_pairwise_delta_correlation.png",
            "06_longitudinal_standardized_change.png",
            "07_within_between_variance.png",
            "08_residual_information.png",
            "09_candidate_fold_consistency.png",
            "10_final_decision_matrix.png",
        )
    )
    required.extend((REPORTS / "table_schema_report.md", REPORTS / "final_report.md"))
    required.extend(
        (
            EXP_ROOT / "EXPERIMENT_PLAN.md",
            Path(__file__).resolve(),
            EXP_ROOT / "scripts" / "verify_outputs.py",
        )
    )
    write_manifest(required)

    print(
        json.dumps(
            {
                "status": "complete",
                "recommended_target": "LD",
                "second_choice": "SPH",
                "input_mismatched": "BPE",
                "formal_patients": len(formal),
                "fold_train_counts": mapping_summary["fold_train_overlap_counts"],
                "outputs": len(required),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
